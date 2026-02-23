import streamlit as st
import re
import urllib.parse
import pandas as pd
import io
import os
import requests
from datetime import datetime

# Sayfa yapılandırması
st.set_page_config(
    page_title="A101 Kampanya Asistanı",
    page_icon="📢",
    layout="wide"
)

# =============================================================================
# CSS STİLLERİ
# =============================================================================
st.markdown("""
<style>
    .main-header {
        text-align: center;
        color: #E31E24;
        font-size: 32px;
        font-weight: bold;
        margin-bottom: 20px;
    }
    .magaza-bandi {
        background-color: #E31E24;
        color: white;
        padding: 20px;
        border-radius: 10px;
        text-align: center;
        font-size: 24px;
        font-weight: bold;
        margin: 20px 0;
        box-shadow: 0 4px 15px rgba(227, 30, 36, 0.3);
    }
    .mesaj-onizleme {
        background-color: #DCF8C6;
        border-radius: 10px;
        padding: 20px;
        font-family: 'Segoe UI', sans-serif;
        font-size: 14px;
        white-space: pre-wrap;
        border-left: 4px solid #25D366;
        margin: 15px 0;
    }
    .uyari-kutusu {
        background-color: #fff3cd;
        border: 1px solid #ffc107;
        border-radius: 8px;
        padding: 15px;
        margin: 10px 0;
    }
    .hata-kutusu {
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        border-radius: 8px;
        padding: 15px;
        margin: 10px 0;
    }
    .basari-kutusu {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 8px;
        padding: 15px;
        margin: 10px 0;
    }
    .secim-rehberi {
        background-color: #e7f3ff;
        border-left: 4px solid #0066cc;
        padding: 15px;
        margin: 15px 0;
        border-radius: 0 8px 8px 0;
    }
    .kontrol-kutusu {
        background-color: #f8f9fa;
        border: 2px solid #dee2e6;
        border-radius: 10px;
        padding: 20px;
        margin: 20px 0;
    }
    .tarih-bilgi {
        background-color: #e8f5e9;
        border-left: 4px solid #4caf50;
        padding: 10px 15px;
        margin: 10px 0;
        border-radius: 0 8px 8px 0;
    }
    .puan-badge {
        display: inline-block;
        padding: 2px 8px;
        border-radius: 12px;
        font-weight: bold;
        font-size: 12px;
        margin-left: 5px;
    }
    .puan-yuksek {
        background-color: #28a745;
        color: white;
    }
    .puan-orta {
        background-color: #ffc107;
        color: black;
    }
    .puan-dusuk {
        background-color: #dc3545;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# MAĞAZA LİSTESİ (WhatsApp Kanalı Kampanya)
# =============================================================================
MAGAZALAR = {
    "0396": "Köroğlu Muratpaşa",
    "1125": "Portakalçiçeği Muratpaşa",
    "1441": "Tonguç",
    "1715": "Çağlayan Muratpaşa",
    "2454": "Bahçelievler Muratpaşa",
    "3812": "Gazi Bulvarı",
    "4282": "Kara Yusuf Kepez",
    "4667": "Orduevi",
    "5490": "Gençlik",
    "6667": "Hastane Cad Kepez",
    "8243": "Güvenlik",
    "8574": "Düden Şelalesi",
    "8878": "Vali Recep Yazıcıoğlu Kepez",
    "8971": "Kayıkent",
    "9395": "Ömer Buyrukçu Cd",
    "B130": "Kamile Çömlekçi",
    "B548": "Hamidiye Muratpaşa",
    "C007": "15 Temmuz Kepez",
    "C241": "Rasih Kaplan Cd Kepez",
    "C346": "Haseki Kepez",
    "C760": "Ermenek",
    "C820": "Kemerağzı Muratpaşa",
    "D483": "Güzeloluk",
    "D587": "Düden Park Muratpaşa",
    "D705": "Molla Gürani",
    "D706": "Suphi Türel Kepez",
    "E046": "Sosyal Güvenlik",
    "E180": "Aydınlıkevler Muratpaşa",
    "E351": "İsmet Gökşen",
    "F296": "Cahit Sıtkı Muratpaşa",
    "F488": "Aşıkveysel",
    "G874": "Mustafa Koç Camii Kepez",
    "H283": "Fabrikalar Kepez",
    "H519": "Üçyol Kepez",
    "H950": "Turgay Koca",
    "I023": "Balbey Muratpaşa",
    "I566": "Nirvana",
    "I693": "Farabi",
    "I824": "Yalı Muratpaşa",
    "J218": "15 Katlılar Kepez",
    "J365": "Kapalı Yol",
    "J433": "IşıkCaddesi",
    "J506": "Yahya Kemal Kepez",
    "J751": "Yeni Niğdeli Parkı",
    "K446": "Paşa Caddesi",
    "K484": "Aydın Kanza",
    "K486": "Müsellim Muratpaşa",
    "K508": "Varlık Muratpaşa",
}

WHATSAPP_NUMBER = "905399311842"

# Performans verisi URL'leri (Google Drive - Yıllık veri)
PERFORMANS_URL_2025 = os.environ.get("PERFORMANS_URL_2025", "")

# =============================================================================
# ÜRÜN EMOJİLERİ
# =============================================================================
URUN_EMOJILERI = {
    # Spesifik olanlar önce (uzun kelimeler) - "ET" içeren kelimeler önce!
    "MASAJ": "💆", "SET": "📦", "SEPET": "🧺", "SEPETİ": "🧺", "KESET": "🧹",
    "KAHVALTILIK": "🥣", "BESLENME KUTUSU": "🍱", "KASE": "🥣",
    "EL ARABASI": "🛒", "BUDAMA": "✂️", "AIRFRYER": "🍟", "POWERBANK": "🔋",
    "SWEATSHIRT": "🧥", "NEVRESİM": "🛏️", "BATTANİYE": "🛏️", "ESPRESSO": "☕",
    "BİSİKLET": "🚲", "VANTİLATÖR": "🌀", "BUZDOLABI": "❄️", "DONDURUC": "🧊",
    "MULTIMEDIA": "🎵", "TESTERE": "🪚", "ÇARŞAF": "🛏️", "SAKLAMA KAB": "📦",
    "MEYVE BIÇAĞI": "🔪", "BIÇAK SETİ": "🔪", "BIÇAK SETI": "🔪",
    "AYAKKABI": "👟", "ÇEKECEĞİ": "🪝", "APARAT": "🔧", "PONPON": "🧶",
    # Meyve ve sebzeler
    "MUZ": "🍌", "PATATES": "🥔", "SOĞAN": "🧅", "DOMATES": "🍅", "ELMA": "🍎",
    "PORTAKAL": "🍊", "LİMON": "🍋", "ÜZÜM": "🍇", "ÇİLEK": "🍓", "KARPUZ": "🍉",
    "HAVUÇ": "🥕", "MARUL": "🥬", "SALATALIK": "🥒", "BİBER": "🌶️", "MANTAR": "🍄",
    "MISIR": "🌽", "BROKOLI": "🥦", "PATLICAN": "🍆", "AVOKADO": "🥑",
    # Tavuk ve et
    "TAVUK": "🍗", "BAGET": "🍗", "BUT": "🍗", "PİLİÇ": "🍗", "KANAT": "🍗",
    "ET": "🥩", "KÖFTE": "🍖", "SUCUK": "🥓", "SOSIS": "🌭",
    # Mutfak eşyaları
    "TAVA": "🍳", "TENCERE": "🍲", "BARDAK": "🥛", "BIÇAK": "🔪", "FİNCAN": "☕",
    "ÇATAL": "🍴", "KAŞIK": "🥄", "TABAK": "🍽️", "KAVANOZ": "🫙", "KEPÇe": "🥄",
    # Tekstil
    "TİŞÖRT": "👕", "KAZAK": "🧥", "BERE": "🧢", "HALI": "🛋️", "KİLİM": "🛋️",
    # Genel kategoriler
    "TV": "📺", "SÜPÜRGE": "🧹", "KLİMA": "❄️",
    "KAHVE": "☕", "ÇAY": "🍵", "TOST": "🥪", "WAFFLE": "🧇",
    "MİKSER": "🥣", "BLENDER": "🥤", "FRİTÖZ": "🍟",
    "SAÇ": "💇", "ÜTÜ": "👔", "ISITICI": "🔥",
    "KAMP": "⛺", "BAHÇE": "🌿", "MANGAL": "🔥", "ŞEMSİYE": "☂️",
    "ARABA": "🚗", "AKÜLÜ": "🔋", "OYUNCAK": "🧸", "BEBEK": "👶",
    "GÖMLEK": "👔", "EŞOFMAN": "🏃", "PANTOLON": "👖", "MONT": "🧥",
    "PERDE": "🪟", "DOLAP": "🗄️", "MASA": "🪑", "SANDALYE": "🪑",
    "TERMOS": "🧊", "SAAT": "⌚", "KAMERA": "📷", "TELEFON": "📱",
    "ÇAPA": "🚜", "MUG": "☕", "VALIZ": "🧳",
    "KUTU": "📦", "RAF": "📚", "AYNA": "🪞", "LAMBA": "💡",
    "DETERJAN": "🧴", "ŞAMPUAN": "🧴", "HAVLU": "🛁", "YORGAN": "🛏️",
    "BOYA": "🎨", "FIRIN": "🔥", "OCAK": "🔥",
    "YASTIK": "🛏️", "PASPAS": "🧹", "POŞET": "🛍️", "ÇÖP": "🗑️",
}

def get_emoji(urun_adi):
    """Ürün adına göre emoji döndür"""
    import re
    urun_upper = str(urun_adi).upper()
    for keyword, emoji in URUN_EMOJILERI.items():
        # "ET" için kelime sınırı kontrolü (METAL, KÜVET, NEHİR içinde eşleşmesin)
        if keyword == "ET":
            if re.search(r'\bET\b', urun_upper):
                return emoji
        elif keyword in urun_upper:
            return emoji
    return "🏷️"

# =============================================================================
# PERFORMANS VERİSİ
# =============================================================================
import os
import tempfile
import gc

# Google Drive File ID - Streamlit secrets veya environment variable'dan al
def _get_gdrive_file_id():
    # Önce Streamlit secrets dene
    try:
        return st.secrets["GDRIVE_FILE_ID"]
    except:
        pass
    # Sonra environment variable
    return os.environ.get("GDRIVE_FILE_ID", "")

@st.cache_resource
def get_perf_local_path() -> str:
    """Parquet'i Google Drive'dan indir - gdown ile otomatik virus scan bypass"""
    import gdown

    file_id = _get_gdrive_file_id()
    if not file_id:
        raise RuntimeError("GDRIVE_FILE_ID secret veya environment variable tanımlı değil!")

    output_path = os.path.join(tempfile.gettempdir(), "veri_yillik.parquet")
    url = f"https://drive.google.com/uc?id={file_id}"
    gdown.download(url, output_path, quiet=False)

    # Parquet signature kontrol (baş PAR1)
    with open(output_path, "rb") as f:
        if f.read(4) != b"PAR1":
            raise RuntimeError("İndirilen dosya parquet değil / bozuk indi (PAR1 yok).")

    return output_path

@st.cache_resource
def _load_parquet_once():
    """
    Parquet dosyasını BİR KERE yükle ve cache'le.
    Tüm fonksiyonlar bu tek kaynaktan okur - RAM tasarrufu!

    RAM Optimizasyonları:
    - Category dtype kullanımı (~%70 RAM tasarrufu)
    - String kolonları category'ye çevir
    - Float32 kullanımı (float64 yerine)
    """
    try:
        path = get_perf_local_path()
        cols = ["Magaza_Kod", "Nitelik", "Urun_Kod", "Satis_Miktari", "Satis_Hasilati_VD", "Mal_Grubu", "Ust_Mal_Grubu"]

        # PyArrow ile oku - daha hafif
        df = pd.read_parquet(path, columns=cols, engine='pyarrow')

        # String kolonları strip'le ve category'ye çevir (RAM tasarrufu!)
        for col in ['Urun_Kod', 'Magaza_Kod', 'Nitelik', 'Mal_Grubu', 'Ust_Mal_Grubu']:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip().astype('category')

        # Float64'ü float32'ye düşür (RAM tasarrufu)
        for col in ['Satis_Miktari', 'Satis_Hasilati_VD']:
            if col in df.columns:
                df[col] = df[col].astype('float32')

        # Garbage collection
        gc.collect()

        return df
    except Exception as e:
        st.warning(f"⚠️ Parquet yüklenemedi: {str(e)}")
        return None

@st.cache_data(ttl=3600)
def load_perf_lookups():
    """Lookup map'lerini oluştur - ana DF'den türetir (RAM dostu)"""
    try:
        df = _load_parquet_once()
        if df is None:
            return {}, {}, []

        # Nitelikler (distinct)
        nitelikler = sorted(df["Nitelik"].dropna().unique().tolist())

        # Urun -> Mal_Grubu (first) - sadece gerekli kolonlarla groupby
        urun_mal_grubu_map = df.groupby("Urun_Kod")["Mal_Grubu"].first().to_dict()
        urun_mal_grubu_map = {str(k).strip(): v for k, v in urun_mal_grubu_map.items() if k is not None}

        # Urun -> Ust_Mal_Grubu (first)
        urun_ust_mal_grubu_map = df.groupby("Urun_Kod")["Ust_Mal_Grubu"].first().to_dict()
        urun_ust_mal_grubu_map = {str(k).strip(): v for k, v in urun_ust_mal_grubu_map.items() if k is not None and v is not None}

        return urun_mal_grubu_map, urun_ust_mal_grubu_map, nitelikler
    except Exception as e:
        st.warning(f"⚠️ Lookup verisi yüklenemedi: {str(e)}")
        return {}, {}, []

@st.cache_data(ttl=3600)
def load_performans_data():
    """Performans DF - ana cache'den döndürür (RAM dostu)"""
    return _load_parquet_once()

@st.cache_data(ttl=3600)
def get_urun_mal_grubu_map(_df):
    """Tüm ürünlerin mal gruplarını döndür (ürün kodu -> mal grubu)"""
    if _df is None:
        return {}
    try:
        urun_mg = _df.groupby('Urun_Kod')['Mal_Grubu'].first().to_dict()
        return {str(k): v for k, v in urun_mg.items()}
    except:
        return {}

@st.cache_data(ttl=3600)
def get_nitelikler(_df):
    """Parquet'teki tüm Nitelik değerlerini döndür"""
    if _df is None:
        return []
    try:
        return sorted(_df['Nitelik'].unique().tolist())
    except:
        return []

@st.cache_data
def build_magaza_options(stok_df_hash, sm_col, bs_col, kod_col, ad_col):
    """Mağaza seçeneklerini cache'le - hızlı filtre için"""
    # Bu fonksiyon stok_df'nin hash'i ile çağrılacak
    return None  # Placeholder - gerçek implementasyon aşağıda

def prepare_magaza_hierarchy(stok_df):
    """SM/BS/Mağaza hiyerarşisini hazırla (vektörel, hızlı)"""
    cols = ['SM', 'BS', 'Kod', 'Mağaza Adı']
    available_cols = [c for c in cols if c in stok_df.columns]

    base = stok_df[available_cols].drop_duplicates().copy()
    base['Kod'] = base['Kod'].astype(str).str.strip()

    if 'SM' in base.columns:
        base['SM'] = base['SM'].astype(str).str.strip()
        sm_list = sorted(base['SM'].dropna().unique().tolist())
    else:
        sm_list = []

    if 'BS' in base.columns:
        base['BS'] = base['BS'].astype(str).str.strip()
        bs_all = sorted(base['BS'].dropna().unique().tolist())
        if 'SM' in base.columns:
            sm_to_bs = (base.groupby('SM')['BS']
                        .apply(lambda x: sorted([b for b in x.unique() if pd.notna(b)]))
                        .to_dict())
        else:
            sm_to_bs = {}
    else:
        bs_all = []
        sm_to_bs = {}

    # Vektörel option oluşturma (iterrows yok!)
    base['opt'] = base['Kod'] + " - " + base['Mağaza Adı'].astype(str)

    return base, sm_list, bs_all, sm_to_bs

@st.cache_data(ttl=3600)
def prepare_lift_aggregations(_performans_df_hash):
    """
    Lift hesaplaması için aggregasyonları önceden hazırla (döngü dışında)
    3 Seviyeli Hiyerarşi: SKU → Mal Grubu → Üst Mal Grubu
    CACHED: Aynı veri için tekrar hesaplamaz - RAM ve CPU tasarrufu!
    """
    performans_df = _load_parquet_once()
    if performans_df is None:
        return None

    # Spot verilerini filtrele - .copy() KULLANMIYORUZ, view ile çalışıyoruz
    spot_mask = performans_df['Nitelik'].str.lower().str.contains('spot', na=False)
    spot_df = performans_df.loc[spot_mask]  # view, copy değil

    # Benchmark toplam
    bench_total = spot_df['Satis_Miktari'].sum()

    # Mağaza bazlı toplamlar
    store_totals = spot_df.groupby('Magaza_Kod')['Satis_Miktari'].sum().to_dict()

    # SKU bazlı - mağaza
    store_sku_qty = spot_df.groupby(['Magaza_Kod', 'Urun_Kod'])['Satis_Miktari'].sum().to_dict()

    # SKU bazlı - benchmark
    bench_sku_qty = spot_df.groupby('Urun_Kod')['Satis_Miktari'].sum().to_dict()

    # Mal grubu bazlı - mağaza
    store_grp_qty = spot_df.groupby(['Magaza_Kod', 'Mal_Grubu'])['Satis_Miktari'].sum().to_dict()

    # Mal grubu bazlı - benchmark
    bench_grp_qty = spot_df.groupby('Mal_Grubu')['Satis_Miktari'].sum().to_dict()

    # Üst mal grubu bazlı - mağaza
    store_ust_grp_qty = spot_df.groupby(['Magaza_Kod', 'Ust_Mal_Grubu'])['Satis_Miktari'].sum().to_dict()

    # Üst mal grubu bazlı - benchmark
    bench_ust_grp_qty = spot_df.groupby('Ust_Mal_Grubu')['Satis_Miktari'].sum().to_dict()

    # Ürün kodu -> Mal grubu ve Üst mal grubu mapping
    urun_mal_grubu = spot_df.groupby('Urun_Kod')['Mal_Grubu'].first().to_dict()
    urun_ust_mal_grubu = spot_df.groupby('Urun_Kod')['Ust_Mal_Grubu'].first().to_dict()

    return {
        'bench_total': bench_total,
        'store_totals': store_totals,
        'store_sku_qty': store_sku_qty,
        'bench_sku_qty': bench_sku_qty,
        'store_grp_qty': store_grp_qty,
        'bench_grp_qty': bench_grp_qty,
        'store_ust_grp_qty': store_ust_grp_qty,
        'bench_ust_grp_qty': bench_ust_grp_qty,
        'urun_mal_grubu': urun_mal_grubu,
        'urun_ust_mal_grubu': urun_ust_mal_grubu
    }


def write_excel_with_formulas(df, sheet_name='Kampanya Önerisi'):
    """
    Excel dosyası oluştur ve yeni marj kolonuna formül ekle.
    Formül: =IF(L{row}="","",((L{row}-((L{row}/100)*(IF(K{row}=1,1,IF(K{row}=5,10,IF(K{row}=6,20,0))))))-H{row})/L{row}*100)

    Kolon sırası:
    A:SM, B:BS, C:Kod, D:Mağaza Adı, E:Ürün Kodu, F:Ürün Tanımı, G:Stok, H:Alış,
    I:Satış Fiyatı, J:Marj, K:kdv, L:yeni fiyat, M:yeni marj, N:Stok TL,
    O:üst mal grubu, P:mal grubu, Q:Lift Skoru, R:Lift, S:Öneri Nedeni
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Kolon sırasını ayarla
    kolon_sirasi = [
        'SM', 'BS', 'Kod', 'Mağaza Adı', 'Ürün Kodu', 'Ürün Tanımı', 'Stok', 'Alış',
        'Satış Fiyatı', 'Marj', 'kdv', 'yeni fiyat', 'yeni marj', 'Stok TL',
        'üst mal grubu', 'mal grubu', 'Lift Skoru', 'Lift', 'Öneri Nedeni'
    ]

    # Mevcut kolonları kontrol et ve sırala
    mevcut_kolonlar = [k for k in kolon_sirasi if k in df.columns]
    df_ordered = df[mevcut_kolonlar].copy()

    # Excel workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header yaz
    for col_idx, col_name in enumerate(df_ordered.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data yaz (formül hariç)
    yeni_marj_col_idx = None
    if 'yeni marj' in df_ordered.columns:
        yeni_marj_col_idx = list(df_ordered.columns).index('yeni marj') + 1

    for row_idx, row in enumerate(df_ordered.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            if col_idx == yeni_marj_col_idx:
                # Formül ekle - yeni marj kolonuna
                # K=kdv (11. kolon), L=yeni fiyat (12. kolon), H=Alış (8. kolon)
                formula = f'=IF(L{row_idx}="","",((L{row_idx}-((L{row_idx}/100)*(IF(K{row_idx}=1,1,IF(K{row_idx}=5,10,IF(K{row_idx}=6,20,0))))))-H{row_idx})/L{row_idx}*100)'
                ws.cell(row=row_idx, column=col_idx, value=formula)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # BytesIO'ya yaz
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


import math

def calculate_lift_scores(kampanya_urunleri, magaza_kodu, nitelik, df, urun_mal_grubu_map, urun_ust_mal_grubu_map=None, weights=None):
    """
    Lift bazlı puanlama algoritması
    - Benchmark: Tüm mağazalar (aynı nitelik)
    - Mağaza payı / Benchmark payı = Lift
    - Shrinkage ile düzeltme
    - Hiyerarşi: Üst Mal Grubu → Mal Grubu → SKU (sadece satışı varsa)
    - weights: (fit, disc, save) ağırlıkları - varsayılan (0.65, 0.25, 0.10)
    """
    if urun_ust_mal_grubu_map is None:
        urun_ust_mal_grubu_map = {}
    if df is None or df.empty:
        return kampanya_urunleri, 0

    # Ağırlıklar
    w_fit, w_disc, w_save = weights if weights else (0.65, 0.25, 0.10)

    eps = 1e-6
    # Bugfix 3: Case-insensitive spot tespiti
    k = 200 if "spot" in str(nitelik).lower() else 500

    # Tip dönüşümleri load_performans_data'da yapıldı - copy() YOK (RAM tasarrufu)
    magaza_kodu_str = str(magaza_kodu).strip()

    # Mağaza ve Benchmark filtreleme (aynı nitelik)
    store_df = df[(df['Magaza_Kod'] == magaza_kodu_str) & (df['Nitelik'] == nitelik)]
    bench_df = df[df['Nitelik'] == nitelik]  # Tüm mağazalar = benchmark

    if store_df.empty:
        # Bu nitelikte mağaza verisi yok, fallback
        return kampanya_urunleri, 0

    # Paydalar (toplam değerler)
    TOTAL_ADET_store = store_df['Satis_Miktari'].sum()
    TOTAL_CIRO_store = store_df['Satis_Hasilati_VD'].sum()
    TOTAL_ADET_bench = bench_df['Satis_Miktari'].sum()
    TOTAL_CIRO_bench = bench_df['Satis_Hasilati_VD'].sum()

    # Bölgedeki mağaza sayısı (ortalama için)
    n_magaza = bench_df['Magaza_Kod'].nunique()
    n_magaza = max(n_magaza, 1)  # Sıfıra bölme önleme

    # Shrinkage weight
    w = TOTAL_ADET_store / (TOTAL_ADET_store + k)

    # === MAL GRUBU LIFT TABLOSU ===
    mal_grubu_lifts = {}
    for g in df['Mal_Grubu'].unique():
        store_g = store_df[store_df['Mal_Grubu'] == g]
        bench_g = bench_df[bench_df['Mal_Grubu'] == g]

        share_qty_store = (store_g['Satis_Miktari'].sum() / TOTAL_ADET_store) if TOTAL_ADET_store > 0 else 0
        share_qty_bench = (bench_g['Satis_Miktari'].sum() / TOTAL_ADET_bench) if TOTAL_ADET_bench > 0 else 0
        share_ciro_store = (store_g['Satis_Hasilati_VD'].sum() / TOTAL_CIRO_store) if TOTAL_CIRO_store > 0 else 0
        share_ciro_bench = (bench_g['Satis_Hasilati_VD'].sum() / TOTAL_CIRO_bench) if TOTAL_CIRO_bench > 0 else 0

        lift_qty = (share_qty_store + eps) / (share_qty_bench + eps)
        lift_ciro = (share_ciro_store + eps) / (share_ciro_bench + eps)

        # Shrinkage
        lift_qty_shr = 1 + w * (lift_qty - 1)
        lift_ciro_shr = 1 + w * (lift_ciro - 1)

        mal_grubu_lifts[g] = {'lift_qty': lift_qty_shr, 'lift_ciro': lift_ciro_shr}

    # === ÜST MAL GRUBU LIFT TABLOSU ===
    ust_mal_grubu_lifts = {}
    if 'Ust_Mal_Grubu' in df.columns:
        for g in df['Ust_Mal_Grubu'].dropna().unique():
            store_g = store_df[store_df['Ust_Mal_Grubu'] == g]
            bench_g = bench_df[bench_df['Ust_Mal_Grubu'] == g]

            share_qty_store = (store_g['Satis_Miktari'].sum() / TOTAL_ADET_store) if TOTAL_ADET_store > 0 else 0
            share_qty_bench = (bench_g['Satis_Miktari'].sum() / TOTAL_ADET_bench) if TOTAL_ADET_bench > 0 else 0
            share_ciro_store = (store_g['Satis_Hasilati_VD'].sum() / TOTAL_CIRO_store) if TOTAL_CIRO_store > 0 else 0
            share_ciro_bench = (bench_g['Satis_Hasilati_VD'].sum() / TOTAL_CIRO_bench) if TOTAL_CIRO_bench > 0 else 0

            lift_qty = (share_qty_store + eps) / (share_qty_bench + eps)
            lift_ciro = (share_ciro_store + eps) / (share_ciro_bench + eps)

            lift_qty_shr = 1 + w * (lift_qty - 1)
            lift_ciro_shr = 1 + w * (lift_ciro - 1)

            ust_mal_grubu_lifts[g] = {'lift_qty': lift_qty_shr, 'lift_ciro': lift_ciro_shr}

    # === SKU LIFT TABLOSU ===
    sku_lifts = {}
    # Key'leri string yap (kampanya ürün kodları string)
    store_sku = {str(k): v for k, v in store_df.groupby('Urun_Kod').agg({'Satis_Miktari': 'sum', 'Satis_Hasilati_VD': 'sum'}).to_dict('index').items()}
    bench_sku = {str(k): v for k, v in bench_df.groupby('Urun_Kod').agg({'Satis_Miktari': 'sum', 'Satis_Hasilati_VD': 'sum'}).to_dict('index').items()}

    for sku in store_sku.keys():
        share_qty_store = (store_sku[sku]['Satis_Miktari'] / TOTAL_ADET_store) if TOTAL_ADET_store > 0 else 0
        share_ciro_store = (store_sku[sku]['Satis_Hasilati_VD'] / TOTAL_CIRO_store) if TOTAL_CIRO_store > 0 else 0

        bench_vals = bench_sku.get(sku, {'Satis_Miktari': 0, 'Satis_Hasilati_VD': 0})
        share_qty_bench = (bench_vals['Satis_Miktari'] / TOTAL_ADET_bench) if TOTAL_ADET_bench > 0 else 0
        share_ciro_bench = (bench_vals['Satis_Hasilati_VD'] / TOTAL_CIRO_bench) if TOTAL_CIRO_bench > 0 else 0

        lift_qty = (share_qty_store + eps) / (share_qty_bench + eps)
        lift_ciro = (share_ciro_store + eps) / (share_ciro_bench + eps)

        lift_qty_shr = 1 + w * (lift_qty - 1)
        lift_ciro_shr = 1 + w * (lift_ciro - 1)

        sku_lifts[str(sku)] = {'lift_qty': lift_qty_shr, 'lift_ciro': lift_ciro_shr}

    # === KAMPANYA ÜRÜNLERİNİ SKORLA ===
    eslesen_sku = 0

    # Güven eşikleri
    SKU_MIN_STORE = 3    # Mağazada min satış adedi
    SKU_MIN_BENCH = 30   # Bölgede min satış adedi
    GROUP_MIN_SHARE = 0.003  # Mal grubu min pay (%0.3)
    ALPHA_K = 5          # Hiyerarşik blend katsayısı

    for urun in kampanya_urunleri:
        urun_kodu = urun.get('kod', '')
        mal_grubu = urun_mal_grubu_map.get(urun_kodu)
        ust_mal_grubu = urun_ust_mal_grubu_map.get(urun_kodu)

        # İndirim skorları
        try:
            eski_fiyat = float(urun.get('eski_fiyat', '0').replace('.', '').replace(',', '.'))
            yeni_fiyat = float(urun.get('yeni_fiyat', '0').replace('.', '').replace(',', '.'))
            saving_tl = eski_fiyat - yeni_fiyat
        except:
            saving_tl = 0

        discount_pct = urun.get('indirim_num', 0) / 100
        disc_score = min(discount_pct / 0.35, 1)  # %35+ = 1
        save_score = math.log1p(saving_tl) / math.log1p(3000) if saving_tl > 0 else 0
        save_score = min(save_score, 1)  # 3000 TL üstü tasarrufta 1'i aşmasın

        # Değişkenler
        fit = 0
        lift_qty = 1
        lift_ciro = 1
        sku_match = False
        store_qty = 0
        store_ciro = 0
        bench_qty = 0
        bench_ciro = 0
        store_share_qty = 0
        store_share_ciro = 0
        bench_share_qty = 0
        bench_share_ciro = 0

        # Uyarı mesajları
        data_warning = None
        group_warning = None
        score_penalty = 0

        # === ÜST MAL GRUBU DEĞERLERİ ===
        fit_ust_group = 0
        lift_qty_ust_group = 1
        lift_ciro_ust_group = 1
        store_ust_group_qty = 0

        if ust_mal_grubu and ust_mal_grubu in ust_mal_grubu_lifts:
            store_ust_g = store_df[store_df['Ust_Mal_Grubu'] == ust_mal_grubu]
            store_ust_group_qty = store_ust_g['Satis_Miktari'].sum()

            lift_qty_ust_group = ust_mal_grubu_lifts[ust_mal_grubu]['lift_qty']
            lift_ciro_ust_group = ust_mal_grubu_lifts[ust_mal_grubu]['lift_ciro']
            fit_ust_group = 0.7 * math.log(max(lift_qty_ust_group, 0.01)) + 0.3 * math.log(max(lift_ciro_ust_group, 0.01))

        # === MAL GRUBU DEĞERLERİ ===
        store_group_qty = 0
        store_group_ciro = 0
        store_group_share = 0
        fit_group = 0
        lift_qty_group = 1
        lift_ciro_group = 1

        if mal_grubu and mal_grubu in mal_grubu_lifts:
            store_g = store_df[store_df['Mal_Grubu'] == mal_grubu]
            bench_g = bench_df[bench_df['Mal_Grubu'] == mal_grubu]
            store_group_qty = store_g['Satis_Miktari'].sum()
            store_group_ciro = store_g['Satis_Hasilati_VD'].sum()
            store_group_share = (store_group_qty / TOTAL_ADET_store) if TOTAL_ADET_store > 0 else 0

            lift_qty_group = mal_grubu_lifts[mal_grubu]['lift_qty']
            lift_ciro_group = mal_grubu_lifts[mal_grubu]['lift_ciro']
            fit_group = 0.7 * math.log(max(lift_qty_group, 0.01)) + 0.3 * math.log(max(lift_ciro_group, 0.01))

        # === MAL GRUBU VARLIK KAPISI ===
        if store_group_qty == 0:
            group_warning = "⛔ Mal grubu hiç satılmamış"
            score_penalty = 50
        elif store_group_share < GROUP_MIN_SHARE:
            group_warning = f"⚠️ Mal grubu zayıf (%{store_group_share*100:.2f})"
            score_penalty = 25

        # === HİYERARŞİK PUANLAMA ===
        # Öncelik: 1) SKU (sadece satışı varsa) 2) Mal Grubu 3) Üst Mal Grubu

        sku_qty_raw = 0
        bench_qty_raw = 0

        # SKU lift SADECE mağazada satış varsa uygulanır (yeni ürünler hariç)
        if urun_kodu in sku_lifts:
            sku_qty_raw = store_sku.get(urun_kodu, {}).get('Satis_Miktari', 0)
            bench_vals = bench_sku.get(urun_kodu, {'Satis_Miktari': 0, 'Satis_Hasilati_VD': 0})
            bench_qty_raw = bench_vals['Satis_Miktari']

        # SKU boost SADECE mağazada satış geçmişi varsa
        has_store_sales = sku_qty_raw > 0

        if has_store_sales and urun_kodu in sku_lifts:
            # Mağazada satış var → SKU lift uygula
            sku_trusted = (sku_qty_raw >= SKU_MIN_STORE) and (bench_qty_raw >= SKU_MIN_BENCH)

            if sku_trusted:
                # SKU verisi güvenilir
                lift_qty = sku_lifts[urun_kodu]['lift_qty']
                lift_ciro = sku_lifts[urun_kodu]['lift_ciro']
                fit_sku = 0.7 * math.log(max(lift_qty, 0.01)) + 0.3 * math.log(max(lift_ciro, 0.01))
                fit = fit_sku
                sku_match = True
                eslesen_sku += 1

                store_qty = sku_qty_raw
                store_ciro = store_sku.get(urun_kodu, {}).get('Satis_Hasilati_VD', 0)
                bench_qty = bench_qty_raw
                bench_ciro = bench_vals['Satis_Hasilati_VD']
            else:
                # SKU verisi yetersiz → Hiyerarşik blend (SKU + Mal Grubu + Üst Mal Grubu)
                alpha_sku = sku_qty_raw / (sku_qty_raw + ALPHA_K)

                lift_qty_sku = sku_lifts[urun_kodu]['lift_qty']
                lift_ciro_sku = sku_lifts[urun_kodu]['lift_ciro']
                fit_sku = 0.7 * math.log(max(lift_qty_sku, 0.01)) + 0.3 * math.log(max(lift_ciro_sku, 0.01))

                # Blend: Mal Grubu ve Üst Mal Grubu ortalaması → SKU ile blend
                fit_category = 0.7 * fit_group + 0.3 * fit_ust_group if fit_group > 0 else fit_ust_group
                lift_qty_category = 0.7 * lift_qty_group + 0.3 * lift_qty_ust_group if fit_group > 0 else lift_qty_ust_group
                lift_ciro_category = 0.7 * lift_ciro_group + 0.3 * lift_ciro_ust_group if fit_group > 0 else lift_ciro_ust_group

                fit = alpha_sku * fit_sku + (1 - alpha_sku) * fit_category
                lift_qty = alpha_sku * lift_qty_sku + (1 - alpha_sku) * lift_qty_category
                lift_ciro = alpha_sku * lift_ciro_sku + (1 - alpha_sku) * lift_ciro_category

                sku_match = True
                eslesen_sku += 1
                data_warning = f"⚠️ Düşük veri ({sku_qty_raw} adet), kategori profili ağırlıklı"

                store_qty = store_group_qty
                store_ciro = store_group_ciro
                bench_qty = bench_g['Satis_Miktari'].sum() if mal_grubu else 0
                bench_ciro = bench_g['Satis_Hasilati_VD'].sum() if mal_grubu else 0

        elif mal_grubu and mal_grubu in mal_grubu_lifts:
            # Yeni ürün veya SKU satışı yok → Mal Grubu + Üst Mal Grubu kullan
            # Ağırlık: %70 Mal Grubu, %30 Üst Mal Grubu
            fit = 0.7 * fit_group + 0.3 * fit_ust_group if fit_ust_group > 0 else fit_group
            lift_qty = 0.7 * lift_qty_group + 0.3 * lift_qty_ust_group if fit_ust_group > 0 else lift_qty_group
            lift_ciro = 0.7 * lift_ciro_group + 0.3 * lift_ciro_ust_group if fit_ust_group > 0 else lift_ciro_group

            store_qty = store_group_qty
            store_ciro = store_group_ciro
            bench_qty = bench_g['Satis_Miktari'].sum()
            bench_ciro = bench_g['Satis_Hasilati_VD'].sum()

        elif ust_mal_grubu and ust_mal_grubu in ust_mal_grubu_lifts:
            # Mal grubu yok ama Üst Mal Grubu var
            fit = fit_ust_group
            lift_qty = lift_qty_ust_group
            lift_ciro = lift_ciro_ust_group

            store_qty = store_ust_group_qty
            store_ciro = 0
            bench_qty = 0
            bench_ciro = 0

        # Pay yüzdeleri
        store_share_qty = (store_qty / TOTAL_ADET_store * 100) if TOTAL_ADET_store > 0 else 0
        store_share_ciro = (store_ciro / TOTAL_CIRO_store * 100) if TOTAL_CIRO_store > 0 else 0
        bench_share_qty = (bench_qty / TOTAL_ADET_bench * 100) if TOTAL_ADET_bench > 0 else 0
        bench_share_ciro = (bench_ciro / TOTAL_CIRO_bench * 100) if TOTAL_CIRO_bench > 0 else 0

        # Final skor: 0.65*fit + 0.25*disc + 0.10*save
        fit_normalized = (fit + 2) / 4  # -2,+2 -> 0,1
        fit_normalized = max(0, min(1, fit_normalized))

        score = w_fit * fit_normalized + w_disc * disc_score + w_save * save_score
        score_100 = round(score * 100, 1)

        # Mal grubu cezası uygula
        score_100 = max(0, score_100 - score_penalty)

        # Sonuçları ürüne ekle
        urun['magaza_skor'] = score_100
        urun['genel_skor'] = round((0.60 * disc_score + 0.25 * fit_normalized + 0.15 * save_score) * 100, 1)
        urun['puan_detay'] = {
            'mal_grubu_adi': mal_grubu or 'Yeni Ürün',
            'lift_qty': round(lift_qty, 2),
            'lift_ciro': round(lift_ciro, 2),
            'disc_score': round(disc_score * 100, 1),
            'save_score': round(save_score * 100, 1),
            'fit': round(fit, 3),
            'sku_match': sku_match,
            # Ham değerler
            'store_qty': round(store_qty),
            'store_ciro': round(store_ciro),
            'bench_qty': round(bench_qty),
            'bench_ciro': round(bench_ciro),
            # Pay yüzdeleri
            'store_share_qty': round(store_share_qty, 2),
            'store_share_ciro': round(store_share_ciro, 2),
            'bench_share_qty': round(bench_share_qty, 2),
            'bench_share_ciro': round(bench_share_ciro, 2),
            # Uyarılar
            'data_warning': data_warning,
            'group_warning': group_warning,
            'sku_qty_raw': sku_qty_raw if urun_kodu in sku_lifts else None
        }

    return kampanya_urunleri, eslesen_sku

def apply_diversity_filter(urunler, max_per_group=2, top_n=10):
    """İlk N öneride aynı mal grubundan max X ürün"""
    sorted_urunler = sorted(urunler, key=lambda x: x.get('magaza_skor', 0), reverse=True)

    result = []
    group_count = {}

    for urun in sorted_urunler:
        mal_grubu = urun.get('puan_detay', {}).get('mal_grubu_adi', 'Yeni Ürün')

        if len(result) < top_n:
            # İlk 10 için çeşitlilik kuralı uygula
            if group_count.get(mal_grubu, 0) < max_per_group:
                result.append(urun)
                group_count[mal_grubu] = group_count.get(mal_grubu, 0) + 1
        else:
            # Geri kalanı direkt ekle
            result.append(urun)

    # Çeşitlilik nedeniyle atlananları sona ekle
    for urun in sorted_urunler:
        if urun not in result:
            result.append(urun)

    return result

def dedupe_similar_products(urunler):
    """
    Benzer ürünleri filtrele - sadece en yüksek skorluyu tut.
    Örnek: "KAHVALTILIK 14 CM TOKIO YEŞİL" ve "KAHVALTILIK 12 CM TOKIO YEŞİL"
    → Sadece yüksek skorlu olan kalır
    """
    import re

    def normalize_name(ad):
        """Ürün adından boyut bilgisini çıkar"""
        ad_upper = str(ad).upper()
        # Boyut patternlerini kaldır: 14 CM, 12,5 CM, 11.5CM, 14CM vb.
        ad_clean = re.sub(r'\d+[,.]?\d*\s*CM', '', ad_upper)
        # Çoklu boşlukları tek boşluğa indir
        ad_clean = re.sub(r'\s+', ' ', ad_clean).strip()
        return ad_clean

    # Normalize edilmiş isme göre grupla
    groups = {}
    for urun in urunler:
        ad = urun.get('ad', '')
        normalized = normalize_name(ad)
        skor = urun.get('magaza_skor', 0)

        if normalized not in groups:
            groups[normalized] = urun
        else:
            # Daha yüksek skorlu olanı tut
            if skor > groups[normalized].get('magaza_skor', 0):
                groups[normalized] = urun

    # Orijinal sıralamayı koru (skora göre)
    result = list(groups.values())
    result.sort(key=lambda x: x.get('magaza_skor', 0), reverse=True)
    return result

def get_puan_badge(puan):
    """Puana göre badge HTML döndür"""
    if puan >= 60:
        return f'<span class="puan-badge puan-yuksek">⭐ {puan}</span>'
    elif puan >= 35:
        return f'<span class="puan-badge puan-orta">📊 {puan}</span>'
    else:
        return f'<span class="puan-badge puan-dusuk">📉 {puan}</span>'

# =============================================================================
# MAİL PARSER - Workflow Formatı
# =============================================================================
def parse_kampanya_maili(mail_text):
    """Workflow kampanya mailini parse et - hem satır hem tablo formatı destekler"""

    result = {
        'baslangic': None,
        'bitis': None,
        'onaylayan': None,
        'urunler': [],
        'hatalar': [],
        'uyarilar': []
    }

    lines = mail_text.strip().split('\n')
    lines = [line.strip() for line in lines if line.strip()]

    # Tarihleri bul
    for i, line in enumerate(lines):
        # Aynı satırda tarih olabilir
        if 'Başlangıç' in line:
            tarih_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', line)
            if tarih_match:
                result['baslangic'] = tarih_match.group(1)
            elif i + 1 < len(lines):
                tarih_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', lines[i + 1])
                if tarih_match:
                    result['baslangic'] = tarih_match.group(1)

        if 'Bitiş' in line:
            tarih_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', line)
            if tarih_match:
                result['bitis'] = tarih_match.group(1)
            elif i + 1 < len(lines):
                tarih_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', lines[i + 1])
                if tarih_match:
                    result['bitis'] = tarih_match.group(1)

        if 'Onaylayan' in line:
            # Aynı satırda isim olabilir
            onay_match = re.search(r'Onaylayan.*?([A-ZÇĞİÖŞÜa-zçğıöşü\s]+)\(', line)
            if onay_match:
                result['onaylayan'] = onay_match.group(1).strip()
            elif i + 1 < len(lines):
                result['onaylayan'] = lines[i + 1]

    # Ürünleri parse et - TAB-SEPARATED TABLO FORMATI
    # Format: Ürün Kodu | Ürün Adı | Satış Fiyatı | Tanıtım Fiyatı | İndirim Oranı
    for line in lines:
        # Tab veya çoklu boşlukla ayrılmış satırlar
        parts = re.split(r'\t+|\s{2,}', line)

        # 8 haneli ürün kodu ile başlayan satır
        if parts and re.match(r'^\d{8}$', parts[0].strip()):
            urun = {
                'kod': parts[0].strip(),
                'ad': '',
                'eski_fiyat': '',
                'yeni_fiyat': '',
                'indirim': '',
                'indirim_num': 0
            }

            # Satırdaki parçaları işle
            for part in parts[1:]:
                part = part.strip()
                if not part:
                    continue

                # Fiyat kontrolü (₺149,00 veya 149,00₺ veya 149,00)
                fiyat_match = re.search(r'₺?\s*([\d.,]+)\s*₺?', part)

                if '%' in part:
                    # İndirim oranı
                    indirim_match = re.search(r'%?\s*([\d.,]+)\s*%?', part)
                    if indirim_match:
                        indirim_str = indirim_match.group(1).replace(',', '.')
                        urun['indirim'] = indirim_str
                        try:
                            urun['indirim_num'] = float(indirim_str)
                        except:
                            pass
                elif '₺' in part and fiyat_match:
                    # Fiyat
                    fiyat_str = fiyat_match.group(1)
                    if not urun['eski_fiyat']:
                        urun['eski_fiyat'] = fiyat_str
                    elif not urun['yeni_fiyat']:
                        urun['yeni_fiyat'] = fiyat_str
                elif not urun['ad'] and not re.match(r'^[\d.,₺%\s]+$', part):
                    # Ürün adı (sayı/fiyat/yüzde içermeyen)
                    urun['ad'] = part

            # Ürün geçerliyse ekle
            if urun['ad'] and (urun['yeni_fiyat'] or urun['eski_fiyat']):
                # Fiyat kontrolü
                if urun['eski_fiyat'] and urun['yeni_fiyat']:
                    try:
                        eski = float(urun['eski_fiyat'].replace('.', '').replace(',', '.'))
                        yeni = float(urun['yeni_fiyat'].replace('.', '').replace(',', '.'))
                        if yeni > eski:
                            result['uyarilar'].append(f"⚠️ {urun['ad'][:30]}: Yeni fiyat eskisinden yüksek!")
                    except:
                        pass
                result['urunler'].append(urun)
                continue

    # Eski format desteği (satır satır ayrılmış)
    if not result['urunler']:
        i = 0
        while i < len(lines):
            line = lines[i]

            # 8 haneli ürün kodu bul (tek başına satırda)
            if re.match(r'^\d{8}$', line):
                urun = {
                    'kod': line,
                    'ad': '',
                    'eski_fiyat': '',
                    'yeni_fiyat': '',
                    'indirim': '',
                    'indirim_num': 0
                }

                # Sonraki satırları oku
                j = i + 1
                while j < len(lines) and j < i + 5:
                    next_line = lines[j]

                    # Fiyat regex
                    fiyat_pattern = r'[₺]?\s*([\d.,]+)\s*(?:₺|TL)?'
                    fiyat_match = re.match(fiyat_pattern, next_line.replace(' ', ''))
                    is_indirim = '%' in next_line

                    if fiyat_match and ('₺' in next_line or 'TL' in next_line):
                        fiyat_str = fiyat_match.group(1)
                        if not urun['eski_fiyat']:
                            urun['eski_fiyat'] = fiyat_str
                        elif not urun['yeni_fiyat']:
                            urun['yeni_fiyat'] = fiyat_str
                    elif is_indirim:
                        indirim_match = re.search(r'([\d.,]+)', next_line)
                        if indirim_match:
                            indirim_str = indirim_match.group(1).replace(',', '.')
                            urun['indirim'] = indirim_str
                            try:
                                urun['indirim_num'] = float(indirim_str)
                            except:
                                pass
                    elif not urun['ad'] and not ('₺' in next_line or 'TL' in next_line or '%' in next_line):
                        urun['ad'] = next_line

                    j += 1

                if urun['ad'] and urun['yeni_fiyat']:
                    result['urunler'].append(urun)

                i = j
            else:
                i += 1

    if not result['baslangic'] or not result['bitis']:
        result['uyarilar'].append("⚠️ Kampanya tarihleri bulunamadı, manuel kontrol edin.")

    if not result['urunler']:
        result['hatalar'].append("🔴 Hiç ürün bulunamadı! Mail formatını kontrol edin.")

    return result

# =============================================================================
# MESAJ FORMATLAMA
# =============================================================================
def format_whatsapp_mesaji(magaza_adi, secili_urunler, bitis_tarihi, toplam_urun_sayisi=None):
    """WhatsApp mesajı oluştur"""
    secili_sayi = len(secili_urunler)

    mesaj = f"🛒 A101 {magaza_adi}\n\n"

    # Başlık - yarına özel
    if toplam_urun_sayisi:
        mesaj += f"🔥 YARINA ÖZEL – {toplam_urun_sayisi} üründe indirim var!\n"
    else:
        mesaj += "🔥 YARINA ÖZEL!\n"

    mesaj += f"⭐ Aşağıdakiler öne çıkan {secili_sayi} fırsat:\n\n"

    # Ürünler
    for urun in secili_urunler:
        emoji = get_emoji(urun['ad'])
        ad_kisa = urun['ad'][:40] if len(urun['ad']) <= 40 else urun['ad'][:37] + "..."

        mesaj += f"{emoji} {ad_kisa}\n"
        mesaj += f"✅ {urun['yeni_fiyat']}₺"
        if urun.get('eski_fiyat'):
            mesaj += f" | Eski: {urun['eski_fiyat']}₺"
        if urun.get('indirim'):
            # İndirim küsuratını kaldır
            indirim_int = int(float(str(urun['indirim']).replace(',', '.')))
            mesaj += f" (%{indirim_int} İNDİRİM)"
        mesaj += "\n\n"

    # Alt bilgi - son gün
    mesaj += f"📅 Son gün: {bitis_tarihi} | 📍 Stoklarla sınırlıdır\n\n"
    mesaj += "Listeden çıkmak için ÇIKIŞ yazın."

    return mesaj

# =============================================================================
# ANA UYGULAMA
# =============================================================================

st.markdown('<p class="main-header">📢 A101 Kampanya Asistanı</p>', unsafe_allow_html=True)

# =============================================================================
# MOD SEÇİMİ
# =============================================================================
mod_secim = st.radio(
    "Ne yapmak istiyorsunuz?",
    options=["📨 Mesaj Oluşturucu", "📊 Kampanya Oluşturucu", "📱 WhatsApp Kanalı Kampanya", "📤 Toplu Mesaj"],
    horizontal=True,
    key="mod_secim",
    help="Mesaj Oluşturucu: Müşteri mesajı oluşturur. Kampanya Oluşturucu: SM/BS/Mağaza seçerek kampanya önerir. WhatsApp Kanalı: Tüm WhatsApp mağazaları için otomatik kampanya önerir. Toplu Mesaj: Excel'den her mağaza için mesaj üretir."
)

st.markdown("---")

# =============================================================================
# MESAJ OLUŞTURUCU MODU
# =============================================================================
if mod_secim == "📨 Mesaj Oluşturucu":
    # =============================================================================
    # ADIM 1: MAĞAZA SEÇİMİ
    # =============================================================================
    st.markdown("### 1️⃣ Mağaza Seçimi")

    magaza_secim = st.selectbox(
        "Mağazanızı seçin:",
        options=[""] + [f"{kod} - {ad}" for kod, ad in MAGAZALAR.items()],
        key="magaza_select"
    )

    if magaza_secim:
        magaza_kodu = magaza_secim.split(" - ")[0]
        magaza_adi = MAGAZALAR[magaza_kodu]

        # Mağaza bandı
        st.markdown(f'''
            <div class="magaza-bandi">
                🏪 {magaza_kodu} - {magaza_adi.upper()}
            </div>
        ''', unsafe_allow_html=True)

        st.info(f"📱 WhatsApp liste adı: **{magaza_kodu}_MUSTERI**")

        # Session state için performans lookups
        if "perf_lookups_loaded" not in st.session_state:
            st.session_state["perf_lookups_loaded"] = False
            st.session_state["urun_mal_grubu_map"] = {}
            st.session_state["urun_ust_mal_grubu_map"] = {}
            st.session_state["nitelikler"] = []

        # Performans lookup'larını butonla yükle (hafif - sadece 3 kolon)
        st.markdown("### 📊 Akıllı Puanlama")

        col_load, col_status = st.columns([1, 2])
        with col_load:
            if st.button("📥 Performans Verisini Yükle", key="btn_load_perf", type="primary"):
                with st.spinner("📊 Performans lookupları hazırlanıyor..."):
                    urun_mal_grubu_map, urun_ust_mal_grubu_map, nitelikler = load_perf_lookups()
                    st.session_state["urun_mal_grubu_map"] = urun_mal_grubu_map
                    st.session_state["urun_ust_mal_grubu_map"] = urun_ust_mal_grubu_map
                    st.session_state["nitelikler"] = nitelikler
                    st.session_state["perf_lookups_loaded"] = len(nitelikler) > 0
                    if st.session_state["perf_lookups_loaded"]:
                        st.rerun()

        perf_loaded = st.session_state["perf_lookups_loaded"]
        urun_mal_grubu_map = st.session_state["urun_mal_grubu_map"]
        urun_ust_mal_grubu_map = st.session_state.get("urun_ust_mal_grubu_map", {})
        nitelikler = st.session_state["nitelikler"]

        with col_status:
            if perf_loaded:
                st.success("✅ Performans verisi yüklendi - Akıllı puanlama aktif!")
            else:
                st.warning("⚠️ Performans verisi yüklenmedi - Sadece indirim bazlı sıralama yapılacak")

        if perf_loaded:

            # Nitelik seçimi
            st.markdown("### 📊 Kampanya Niteliği")
            # Varsayılan olarak Spot seçili gelsin
            default_nitelik = "Spot" if "Spot" in nitelikler else (nitelikler[0] if nitelikler else None)
            default_index = nitelikler.index(default_nitelik) if default_nitelik in nitelikler else 0

            nitelik_secim = st.selectbox(
                "Kampanya niteliğini seçin:",
                options=nitelikler,
                index=default_index,
                key="nitelik_select",
                help="Kampanya türüne göre seçin. Genellikle 'Spot' kullanılır."
            )

            # Skor ağırlıkları (gelişmiş ayarlar)
            with st.expander("⚙️ Skor Ağırlıkları (Gelişmiş)"):
                st.caption("Varsayılan değerler önerilir. Değiştirmek isterseniz ayarlayın.")
                col_w1, col_w2, col_w3 = st.columns(3)
                with col_w1:
                    weight_fit = st.slider("Müşteri Uyumu", 0, 100, 65, 5, key="w_fit", help="Lift bazlı uyum skoru") / 100
                with col_w2:
                    weight_disc = st.slider("İndirim", 0, 100, 25, 5, key="w_disc", help="İndirim oranı skoru") / 100
                with col_w3:
                    weight_save = st.slider("Tasarruf", 0, 100, 10, 5, key="w_save", help="TL bazlı tasarruf") / 100

                # Normalize et (toplamı 1 yap)
                total_weight = weight_fit + weight_disc + weight_save
                if total_weight > 0:
                    weight_fit = weight_fit / total_weight
                    weight_disc = weight_disc / total_weight
                    weight_save = weight_save / total_weight
                st.caption(f"Normalize: Uyum {weight_fit:.0%} | İndirim {weight_disc:.0%} | Tasarruf {weight_save:.0%}")
        else:
            nitelik_secim = None

        st.markdown("---")

        # =============================================================================
        # ADIM 2: KAMPANYA MAİLİ YAPIŞTIR
        # =============================================================================
        st.markdown("### 2️⃣ Kampanya Mailini Yapıştırın")

        st.markdown("""
        <div class="secim-rehberi">
            <strong>📋 Nasıl yapılır:</strong><br>
            1. Workflow'dan gelen kampanya onay mailini açın<br>
            2. <strong>Ctrl+A</strong> (tümünü seç) → <strong>Ctrl+C</strong> (kopyala)<br>
            3. Aşağıdaki alana <strong>Ctrl+V</strong> (yapıştır)
        </div>
        """, unsafe_allow_html=True)

        mail_icerik = st.text_area(
            "Kampanya mailini buraya yapıştırın:",
            height=200,
            placeholder="Mağaza Bölgesel Tanıtım Sonucu\n\nTanıtım Başlangıç Tarihi\n20.12.2025\n..."
        )

        if mail_icerik:
            # Parse et
            kampanya = parse_kampanya_maili(mail_icerik)

            # Hataları göster
            if kampanya['hatalar']:
                for hata in kampanya['hatalar']:
                    st.error(hata)
                st.stop()

            # Ürünleri puanla (Lift bazlı algoritma)
            if nitelik_secim and perf_loaded:
                # Full performans DF'yi yükle (lift hesabı için gerekli)
                performans_df = load_performans_data()

                if performans_df is not None:
                    # Ağırlıkları session_state'den al
                    w_fit = st.session_state.get('w_fit', 65) / 100
                    w_disc = st.session_state.get('w_disc', 25) / 100
                    w_save = st.session_state.get('w_save', 10) / 100
                    # Normalize et
                    total_w = w_fit + w_disc + w_save
                    if total_w > 0:
                        weights = (w_fit/total_w, w_disc/total_w, w_save/total_w)
                    else:
                        weights = (0.65, 0.25, 0.10)

                    kampanya['urunler'], eslesen_sku = calculate_lift_scores(
                        kampanya['urunler'],
                        magaza_kodu,
                        nitelik_secim,
                        performans_df,
                        urun_mal_grubu_map,
                        urun_ust_mal_grubu_map=urun_ust_mal_grubu_map,
                        weights=weights
                    )
                else:
                    eslesen_sku = 0
                    for urun in kampanya['urunler']:
                        disc = urun.get('indirim_num', 0) / 100
                        urun['magaza_skor'] = round(min(disc / 0.35, 1) * 100, 1)
                        urun['genel_skor'] = urun['magaza_skor']
                        urun['puan_detay'] = {'mal_grubu_adi': urun_mal_grubu_map.get(urun.get('kod', ''), 'Yeni Ürün')}
            else:
                eslesen_sku = 0
                # Fallback: sadece indirim bazlı
                for urun in kampanya['urunler']:
                    disc = urun.get('indirim_num', 0) / 100
                    urun['magaza_skor'] = round(min(disc / 0.35, 1) * 100, 1)
                    urun['genel_skor'] = urun['magaza_skor']
                    urun['puan_detay'] = {'mal_grubu_adi': urun_mal_grubu_map.get(urun.get('kod', ''), 'Yeni Ürün')}

            # Eşleşme sayısını hesapla
            toplam_urun = len(kampanya['urunler'])
            eslesen_mg = sum(1 for u in kampanya['urunler']
                              if u.get('puan_detay', {}).get('mal_grubu_adi')
                              and u.get('puan_detay', {}).get('mal_grubu_adi') != 'Yeni Ürün')

            # Başarı mesajı
            st.markdown(f'''
                <div class="basari-kutusu">
                    <strong>✅ {toplam_urun} ürün okundu ve puanlandı</strong><br>
                    📊 SKU eşleşmesi: <strong>{eslesen_sku}/{toplam_urun}</strong> |
                    Mal grubu eşleşmesi: <strong>{eslesen_mg}/{toplam_urun}</strong>
                </div>
            ''', unsafe_allow_html=True)

            # Tarih bilgisi
            if kampanya['baslangic'] and kampanya['bitis']:
                st.markdown(f'''
                    <div class="tarih-bilgi">
                        📅 <strong>Kampanya:</strong> {kampanya['baslangic']} - {kampanya['bitis']}
                    </div>
                ''', unsafe_allow_html=True)

            # Uyarıları göster
            if kampanya['uyarilar']:
                with st.expander(f"⚠️ {len(kampanya['uyarilar'])} Uyarı", expanded=False):
                    for uyari in kampanya['uyarilar']:
                        st.warning(uyari)

            st.markdown("---")

            # =============================================================================
            # ADIM 3: ÜRÜN SEÇİMİ (İKİ SIRALAMA)
            # =============================================================================
            st.markdown("### 3️⃣ Ürün Seçimi")

            # En iyi 5 öneri butonu
            col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
            with col_btn1:
                if st.button("🎯 En İyi 5 Öneri", type="primary", use_container_width=True):
                    # Önce benzer ürünleri filtrele, sonra çeşitlilik uygula
                    deduped = dedupe_similar_products(kampanya['urunler'])
                    top5 = apply_diversity_filter(deduped, max_per_group=2, top_n=5)
                    st.session_state['auto_selected'] = [u['kod'] for u in top5]
                    st.rerun()
            with col_btn2:
                if st.button("🔄 Seçimleri Temizle", use_container_width=True):
                    st.session_state['auto_selected'] = []
                    st.rerun()

            # Otomatik seçim listesi
            auto_selected = st.session_state.get('auto_selected', [])

            # İki tab ile iki farklı sıralama
            tab_magaza, tab_genel = st.tabs([
                f"🏪 {magaza_adi} İçin Önerilen",
                "📊 Genel Öneri (İndirim Bazlı)"
            ])

            secili_urunler = []

            with tab_magaza:
                st.markdown(f"""
                <div class="secim-rehberi">
                    <strong>🏪 Mağaza Bazlı Puanlama (Lift Algoritması):</strong><br>
                    Nitelik: <strong>{nitelik_secim or '-'}</strong> | Benchmark: Tüm Mağazalar<br>
                    • Müşteri Uyumu (65%) - Lift: Mağaza payı / Bölge payı<br>
                    • İndirim Çekiciliği (25%) - %35+ = maksimum<br>
                    • Tasarruf (10%) - TL bazlı (log normalize)<br><br>
                    🟢 60+ Çok İyi | 🟡 35-60 Orta | 🔴 35- Düşük
                </div>
                """, unsafe_allow_html=True)

                # Önce benzer ürünleri filtrele, sonra çeşitlilik uygula
                deduped_urunler = dedupe_similar_products(kampanya['urunler'])
                urunler_magaza = apply_diversity_filter(deduped_urunler, max_per_group=2, top_n=10)

                # Güvenilir ve düşük güvenli ürünleri ayır
                urunler_guvenilir = [u for u in urunler_magaza if not u.get('puan_detay', {}).get('group_warning')]
                urunler_dusuk_guven = [u for u in urunler_magaza if u.get('puan_detay', {}).get('group_warning')]

                # Önce güvenilir ürünler
                for urun in urunler_guvenilir:
                    col1, col2, col3 = st.columns([1, 17, 4])

                    with col1:
                        default_val = urun['kod'] in auto_selected
                        secili = st.checkbox("", key=f"m_{urun['kod']}", value=default_val, label_visibility="collapsed")
                        if secili and urun not in secili_urunler:
                            secili_urunler.append(urun)

                    with col2:
                        emoji = get_emoji(urun['ad'])
                        puan = urun.get('magaza_skor', 0)
                        puan_badge = get_puan_badge(puan)
                        detay = urun.get('puan_detay', {})
                        mal_grubu = detay.get('mal_grubu_adi', '-')
                        sku_icon = "🎯" if detay.get('sku_match') else ""
                        st.markdown(
                            f"{emoji} **{urun['ad'][:40]}** | _{mal_grubu}_ {sku_icon} → {urun['yeni_fiyat']}₺ ~~{urun['eski_fiyat']}₺~~ {puan_badge}",
                            unsafe_allow_html=True
                        )

                    with col3:
                        with st.popover("📊 Detay"):
                            st.write(f"**Mal Grubu:** {mal_grubu}")
                            # Uyarılar
                            if detay.get('group_warning'):
                                st.warning(detay.get('group_warning'))
                            if detay.get('data_warning'):
                                st.info(detay.get('data_warning'))
                            st.markdown("---")
                            st.write("**📦 ADET**")
                            sku_raw = detay.get('sku_qty_raw')
                            if sku_raw is not None:
                                st.write(f"SKU Satış: {sku_raw} adet")
                            st.write(f"Mağaza: {detay.get('store_qty', 0):,} | Pay: %{detay.get('store_share_qty', 0):.2f}")
                            st.write(f"Bölge: {detay.get('bench_qty', 0):,} | Pay: %{detay.get('bench_share_qty', 0):.2f}")
                            st.write(f"**Lift: {detay.get('lift_qty', 1):.2f}x**")
                            st.markdown("---")
                            st.write("**💰 CİRO**")
                            st.write(f"Mağaza: {detay.get('store_ciro', 0):,}₺ | Pay: %{detay.get('store_share_ciro', 0):.2f}")
                            st.write(f"Bölge: {detay.get('bench_ciro', 0):,}₺ | Pay: %{detay.get('bench_share_ciro', 0):.2f}")
                            st.write(f"**Lift: {detay.get('lift_ciro', 1):.2f}x**")
                            st.markdown("---")
                            st.write(f"🏷️ İndirim: {detay.get('disc_score', 0):.0f} | 💵 Tasarruf: {detay.get('save_score', 0):.0f}")
                            st.write(f"🔍 SKU Eşleşme: {'✅' if detay.get('sku_match') else '❌'}")
                            st.markdown("---")
                            st.caption("ℹ️ Lift = Mağaza payı / Bölge payı")
                            st.caption("SKU az satıldıysa grup profili ağırlıklı hesaplanır")

                # Düşük güvenli ürünler (varsa)
                if urunler_dusuk_guven:
                    with st.expander(f"⚠️ Düşük Güvenli Ürünler ({len(urunler_dusuk_guven)} adet)", expanded=False):
                        st.caption("Bu ürünler mağazada zayıf kategorilerden. Dikkatli değerlendirin.")
                        for urun in urunler_dusuk_guven:
                            col1, col2, col3 = st.columns([1, 17, 4])
                            with col1:
                                default_val = urun['kod'] in auto_selected
                                secili = st.checkbox("", key=f"ml_{urun['kod']}", value=default_val, label_visibility="collapsed")
                                if secili and urun not in secili_urunler:
                                    secili_urunler.append(urun)
                            with col2:
                                emoji = get_emoji(urun['ad'])
                                puan = urun.get('magaza_skor', 0)
                                puan_badge = get_puan_badge(puan)
                                detay = urun.get('puan_detay', {})
                                mal_grubu = detay.get('mal_grubu_adi', '-')
                                warning = detay.get('group_warning', '')
                                st.markdown(
                                    f"⚠️ {emoji} **{urun['ad'][:35]}** | _{mal_grubu}_ → {urun['yeni_fiyat']}₺ {puan_badge}",
                                    unsafe_allow_html=True
                                )
                                st.caption(warning)

            with tab_genel:
                st.markdown("""
                <div class="secim-rehberi">
                    <strong>📊 Genel Puanlama (İndirim Ağırlıklı):</strong><br>
                    Bu sıralama <strong>indirim oranına</strong> göre yapılmıştır.<br>
                    • İndirim Oranı (60%)<br>
                    • Müşteri Uyumu (25%)<br>
                    • Tasarruf (15%)<br><br>
                    🟢 60+ Çok İyi | 🟡 35-60 Orta | 🔴 35- Düşük
                </div>
                """, unsafe_allow_html=True)

                # Benzer ürünleri filtrele ve genel skora göre sırala
                urunler_genel = dedupe_similar_products(kampanya['urunler'])
                urunler_genel = sorted(urunler_genel, key=lambda x: x.get('genel_skor', 0), reverse=True)

                for urun in urunler_genel:
                    col1, col2, col3 = st.columns([1, 17, 4])

                    with col1:
                        default_val = urun['kod'] in auto_selected
                        secili = st.checkbox("", key=f"g_{urun['kod']}", value=default_val, label_visibility="collapsed")
                        if secili and urun not in secili_urunler:
                            secili_urunler.append(urun)

                    with col2:
                        emoji = get_emoji(urun['ad'])
                        puan = urun.get('genel_skor', 0)
                        puan_badge = get_puan_badge(puan)
                        detay = urun.get('puan_detay', {})
                        mal_grubu = detay.get('mal_grubu_adi', '-')
                        st.markdown(
                            f"{emoji} **{urun['ad'][:40]}** | _{mal_grubu}_ → {urun['yeni_fiyat']}₺ ~~{urun['eski_fiyat']}₺~~ | %{urun['indirim']} {puan_badge}",
                            unsafe_allow_html=True
                        )

                    with col3:
                        with st.popover("📊 Detay"):
                            st.write(f"**Mal Grubu:** {mal_grubu}")
                            # Uyarılar
                            if detay.get('group_warning'):
                                st.warning(detay.get('group_warning'))
                            if detay.get('data_warning'):
                                st.info(detay.get('data_warning'))
                            st.markdown("---")
                            st.write("**📦 ADET**")
                            sku_raw = detay.get('sku_qty_raw')
                            if sku_raw is not None:
                                st.write(f"SKU Satış: {sku_raw} adet")
                            st.write(f"Mağaza: {detay.get('store_qty', 0):,} | Pay: %{detay.get('store_share_qty', 0):.2f}")
                            st.write(f"Bölge: {detay.get('bench_qty', 0):,} | Pay: %{detay.get('bench_share_qty', 0):.2f}")
                            st.write(f"**Lift: {detay.get('lift_qty', 1):.2f}x**")
                            st.markdown("---")
                            st.write("**💰 CİRO**")
                            st.write(f"Mağaza: {detay.get('store_ciro', 0):,}₺ | Pay: %{detay.get('store_share_ciro', 0):.2f}")
                            st.write(f"Bölge: {detay.get('bench_ciro', 0):,}₺ | Pay: %{detay.get('bench_share_ciro', 0):.2f}")
                            st.write(f"**Lift: {detay.get('lift_ciro', 1):.2f}x**")
                            st.markdown("---")
                            st.write(f"🏷️ İndirim: {detay.get('disc_score', 0):.0f} | 💵 Tasarruf: {detay.get('save_score', 0):.0f}")
                            st.write(f"🔍 SKU Eşleşme: {'✅' if detay.get('sku_match') else '❌'}")
                            st.markdown("---")
                            st.caption("ℹ️ Lift = Mağaza payı / Bölge payı")
                            st.caption("SKU az satıldıysa grup profili ağırlıklı hesaplanır")

            # Seçim kontrolü
            secili_sayi = len(secili_urunler)

            if secili_sayi > 0:
                if secili_sayi < 3:
                    st.warning(f"⚠️ {secili_sayi} ürün seçildi. En az 3 ürün önerilir.")
                elif secili_sayi > 5:
                    st.warning(f"⚠️ {secili_sayi} ürün seçildi. En fazla 5 ürün önerilir.")
                else:
                    st.success(f"✅ {secili_sayi} ürün seçildi")

                st.markdown("---")

                # =============================================================================
                # ADIM 4: STOK KONTROLÜ
                # =============================================================================
                st.markdown("### 4️⃣ Stok Kontrolü")

                stok_onay = st.checkbox(
                    f"✅ Seçtiğim {secili_sayi} ürün **{magaza_adi}** mağazasında STOKTA VAR",
                    key="stok_onay"
                )

                st.markdown("---")

                # =============================================================================
                # ADIM 5: MESAJ ÖNİZLEME VE GÖNDERME
                # =============================================================================
                st.markdown("### 5️⃣ Mesaj Önizleme ve Gönderme")

                # Mesajı oluştur
                bitis = kampanya['bitis'] or "Stoklarla sınırlı"
                toplam_urun = len(kampanya['urunler'])
                mesaj = format_whatsapp_mesaji(magaza_adi, secili_urunler, bitis, toplam_urun)

                st.markdown("**Mesaj önizleme:**")
                st.markdown(f'<div class="mesaj-onizleme">{mesaj}</div>', unsafe_allow_html=True)

                # Kopyala ve Düzenle butonları
                col_kopyala, col_duzenle = st.columns(2)
                with col_kopyala:
                    if st.button("📋 Mesajı Kopyala", type="primary", use_container_width=True):
                        st.session_state['mesaj_kopyala_goster'] = True
                        st.session_state['mesaj_duzenle_goster'] = False
                with col_duzenle:
                    if st.button("✏️ Mesajı Düzenle", use_container_width=True):
                        st.session_state['mesaj_duzenle_goster'] = True
                        st.session_state['mesaj_kopyala_goster'] = False

                # Kopyala modu
                if st.session_state.get('mesaj_kopyala_goster', False):
                    st.code(mesaj, language=None)
                    st.caption("👆 Sağ üst köşedeki 📋 ikonuna tıklayarak kopyalayın")

                # Düzenle modu
                if st.session_state.get('mesaj_duzenle_goster', False):
                    mesaj_duzenle = st.text_area(
                        "Mesajı düzenleyin:",
                        value=mesaj,
                        height=300,
                        key="mesaj_duzenle_area"
                    )
                    if mesaj_duzenle != mesaj:
                        mesaj = mesaj_duzenle
                    st.caption("💡 Düzenledikten sonra Ctrl+A → Ctrl+C ile kopyalayın")

                # Kontroller
                st.markdown("---")
                st.markdown('<div class="kontrol-kutusu">', unsafe_allow_html=True)
                st.markdown("### ⚠️ Gönderim Öncesi Kontrol")

                kontrol1 = st.checkbox(
                    f"✅ Bu mesaj **{magaza_kodu} - {magaza_adi}** için hazırlandı",
                    key="kontrol1"
                )

                kontrol2 = st.checkbox(
                    f"✅ Tarih ({bitis}) ve fiyatlar doğru",
                    key="kontrol2"
                )

                st.markdown('</div>', unsafe_allow_html=True)

                # WhatsApp butonu
                if stok_onay and kontrol1 and kontrol2:
                    encoded_mesaj = urllib.parse.quote(mesaj)
                    whatsapp_link = f"https://wa.me/{WHATSAPP_NUMBER}?text={encoded_mesaj}"

                    st.markdown(f'''
                        <a href="{whatsapp_link}" target="_blank" style="
                            display: block;
                            background-color: #25D366;
                            color: white;
                            padding: 20px 40px;
                            text-decoration: none;
                            border-radius: 10px;
                            font-size: 20px;
                            font-weight: bold;
                            text-align: center;
                            margin-top: 20px;
                            box-shadow: 0 4px 15px rgba(37, 211, 102, 0.4);
                        ">
                            💬 WhatsApp'ta Gönder
                        </a>
                    ''', unsafe_allow_html=True)

                    st.info(f"👆 Butona tıklayınca WhatsApp açılacak. **{magaza_kodu}_MUSTERI** listesini seçip gönderin.")
                else:
                    st.markdown('''
                        <div style="
                            display: block;
                            background-color: #ccc;
                            color: #666;
                            padding: 20px 40px;
                            border-radius: 10px;
                            font-size: 20px;
                            font-weight: bold;
                            text-align: center;
                            margin-top: 20px;
                        ">
                            💬 WhatsApp'ta Gönder
                        </div>
                    ''', unsafe_allow_html=True)
                    st.warning("☝️ Yukarıdaki tüm kontrolleri tamamlayın.")

    else:
        st.info("👆 Önce mağazanızı seçin.")

# =============================================================================
# KAMPANYA OLUŞTURUCU MODU
# =============================================================================
elif mod_secim == "📊 Kampanya Oluşturucu":
    st.markdown("""
    <div class="secim-rehberi">
        <strong>📊 Kampanya Oluşturucu Nasıl Çalışır?</strong><br>
        1. Stok Excel dosyasını yükleyin<br>
        2. SM/BS/Mağaza filtreleri ile hedef mağazaları seçin<br>
        3. Sistem, lift algoritmasıyla hangi ürünlere kampanya yapılmalı önerir<br>
        4. Önerilen ürünleri Excel olarak indirin ve fiyatları belirleyin
    </div>
    """, unsafe_allow_html=True)

    # =============================================================================
    # ADIM 1: EXCEL YÜKLEME
    # =============================================================================
    st.markdown("### 1️⃣ Stok Verisini Yükleyin")

    uploaded_file = st.file_uploader(
        "Stok Excel dosyasını yükleyin (.xlsx)",
        type=['xlsx'],
        key="stok_excel_upload",
        help="Kolonlar: SM, BS, Kod, Mağaza Adı, Ürün Kodu, Ürün Tanımı, Stok, Alış, Satış Fiyatı, Marj, KDV, yeni fiyat, yeni marj, Stok TL"
    )

    if uploaded_file is not None:
        try:
            # Excel'i oku
            stok_df = pd.read_excel(uploaded_file)

            # Kolon isimlerini normalize et (boşlukları temizle)
            stok_df.columns = stok_df.columns.str.strip()

            # Kolon isimlerini standartlaştır (case-insensitive)
            kolon_mapping = {
                'üst mal grubu': 'Üst Mal Grubu',
                'mal grubu': 'Mal Grubu',
                'ürün kodu': 'Ürün Kodu',
                'ürün tanımı': 'Ürün Tanımı',
                'mağaza adı': 'Mağaza Adı',
                'satış fiyatı': 'Satış Fiyatı',
                'kod': 'Kod',
                'stok': 'Stok'
            }
            stok_df.columns = [kolon_mapping.get(col.lower(), col) for col in stok_df.columns]

            # Kod kolonunu string'e çevir (pyarrow mixed type hatası için)
            if 'Kod' in stok_df.columns:
                stok_df['Kod'] = stok_df['Kod'].astype(str).str.strip()

            # Gerekli kolonları kontrol et
            gerekli_kolonlar = ['Kod', 'Mağaza Adı', 'Ürün Kodu', 'Ürün Tanımı', 'Stok', 'Satış Fiyatı', 'Üst Mal Grubu', 'Mal Grubu']
            eksik_kolonlar = [k for k in gerekli_kolonlar if k not in stok_df.columns]

            if eksik_kolonlar:
                st.error(f"❌ Eksik kolonlar: {', '.join(eksik_kolonlar)}")
                st.info("Beklenen kolonlar: SM, BS, Kod, Mağaza Adı, Ürün Kodu, Ürün Tanımı, Stok, Alış, Satış Fiyatı, Marj, KDV, **Üst Mal Grubu**, **Mal Grubu**...")
            else:
                st.success(f"✅ {len(stok_df):,} satır yüklendi")

                # Önizleme
                with st.expander("📋 Veri Önizleme (ilk 10 satır)"):
                    st.dataframe(stok_df.head(10))

                st.markdown("---")

                # =============================================================================
                # ADIM 2: FİLTRELEME (SM → BS → MAĞAZA) - OPTİMİZE v3
                # =============================================================================
                st.markdown("### 2️⃣ Mağaza Filtresi")

                # Hiyerarşiyi 1 KERE hesapla, session_state'e kaydet
                stok_hash = hash(tuple(stok_df['Kod'].astype(str).head(100)))
                if st.session_state.get('stok_hash') != stok_hash:
                    # Sadece gerekli kolonlarla küçük df
                    small_cols = [c for c in ['SM', 'BS', 'Kod', 'Mağaza Adı'] if c in stok_df.columns]
                    small = stok_df[small_cols].drop_duplicates().copy()
                    small['Kod'] = small['Kod'].astype(str).str.strip()
                    if 'SM' in small.columns:
                        small['SM'] = small['SM'].astype(str).str.strip()
                    if 'BS' in small.columns:
                        small['BS'] = small['BS'].astype(str).str.strip()
                    small['opt'] = small['Kod'] + " - " + small['Mağaza Adı'].astype(str)

                    # Lookup map'ler oluştur (1 kere)
                    sm_list = sorted(small['SM'].dropna().unique().tolist()) if 'SM' in small.columns else []
                    bs_all = sorted(small['BS'].dropna().unique().tolist()) if 'BS' in small.columns else []

                    sm_to_bs = {}
                    sm_to_opt = {}
                    bs_to_opt = {}
                    smbs_to_opt = {}
                    all_opts = sorted(small['opt'].unique().tolist())

                    if 'SM' in small.columns and 'BS' in small.columns:
                        sm_to_bs = small.groupby('SM')['BS'].apply(lambda x: sorted(x.dropna().unique().tolist())).to_dict()
                        sm_to_opt = small.groupby('SM')['opt'].apply(lambda x: sorted(x.unique().tolist())).to_dict()
                        bs_to_opt = small.groupby('BS')['opt'].apply(lambda x: sorted(x.unique().tolist())).to_dict()
                        smbs_to_opt = small.groupby(['SM', 'BS'])['opt'].apply(lambda x: sorted(x.unique().tolist())).to_dict()
                    elif 'SM' in small.columns:
                        sm_to_opt = small.groupby('SM')['opt'].apply(lambda x: sorted(x.unique().tolist())).to_dict()
                    elif 'BS' in small.columns:
                        bs_to_opt = small.groupby('BS')['opt'].apply(lambda x: sorted(x.unique().tolist())).to_dict()

                    # Session state'e kaydet
                    st.session_state['stok_hash'] = stok_hash
                    st.session_state['filter_sm_list'] = sm_list
                    st.session_state['filter_bs_all'] = bs_all
                    st.session_state['filter_sm_to_bs'] = sm_to_bs
                    st.session_state['filter_sm_to_opt'] = sm_to_opt
                    st.session_state['filter_bs_to_opt'] = bs_to_opt
                    st.session_state['filter_smbs_to_opt'] = smbs_to_opt
                    st.session_state['filter_all_opts'] = all_opts
                    st.session_state['kampanya_stok_df'] = stok_df

                # Session state'den oku (hızlı!)
                sm_list = st.session_state.get('filter_sm_list', [])
                bs_all = st.session_state.get('filter_bs_all', [])
                sm_to_bs = st.session_state.get('filter_sm_to_bs', {})
                sm_to_opt = st.session_state.get('filter_sm_to_opt', {})
                bs_to_opt = st.session_state.get('filter_bs_to_opt', {})
                smbs_to_opt = st.session_state.get('filter_smbs_to_opt', {})
                all_opts = st.session_state.get('filter_all_opts', [])

                # FORM YOK - Anlık güncelleme için session_state kullan
                col_sm, col_bs = st.columns(2)

                with col_sm:
                    secili_sm = st.multiselect(
                        "SM Seçin (opsiyonel):",
                        options=sm_list,
                        default=st.session_state.get('kamp_secili_sm', []),
                        key="kamp_sm_select",
                        help="Boş bırakırsanız tüm SM'ler dahil edilir"
                    )
                    # Seçim değiştiyse kaydet
                    if secili_sm != st.session_state.get('kamp_secili_sm', []):
                        st.session_state['kamp_secili_sm'] = secili_sm
                        st.session_state['kamp_secili_bs'] = []  # BS sıfırla
                        st.session_state['kamp_secili_magazalar'] = []  # Mağaza sıfırla

                # BS listesini SM'e göre getir (dict lookup - anlık!)
                secili_sm_state = st.session_state.get('kamp_secili_sm', [])
                if secili_sm_state:
                    bs_listesi = sorted(set().union(*[set(sm_to_bs.get(sm, [])) for sm in secili_sm_state]))
                else:
                    bs_listesi = bs_all

                with col_bs:
                    secili_bs = st.multiselect(
                        "BS Seçin (opsiyonel):",
                        options=bs_listesi,
                        default=[b for b in st.session_state.get('kamp_secili_bs', []) if b in bs_listesi],
                        key="kamp_bs_select",
                        help="Boş bırakırsanız tüm BS'ler dahil edilir"
                    )
                    if secili_bs != st.session_state.get('kamp_secili_bs', []):
                        st.session_state['kamp_secili_bs'] = secili_bs
                        st.session_state['kamp_secili_magazalar'] = []  # Mağaza sıfırla

                # Mağaza listesini dict lookup ile getir (anlık!)
                secili_bs_state = st.session_state.get('kamp_secili_bs', [])
                if secili_sm_state and secili_bs_state:
                    magaza_options = sorted(set().union(*[
                        set(smbs_to_opt.get((sm, bs), []))
                        for sm in secili_sm_state for bs in secili_bs_state
                    ]))
                elif secili_sm_state:
                    magaza_options = sorted(set().union(*[set(sm_to_opt.get(sm, [])) for sm in secili_sm_state]))
                elif secili_bs_state:
                    magaza_options = sorted(set().union(*[set(bs_to_opt.get(bs, [])) for bs in secili_bs_state]))
                else:
                    magaza_options = all_opts

                # Hızlı seçim: Mağaza kodlarını yapıştır
                with st.expander("📋 Hızlı Seçim: Mağaza Kodlarını Yapıştır", expanded=False):
                    yapistir_text = st.text_area(
                        "Mağaza kodlarını yapıştırın (her satıra bir kod veya virgülle ayırın):",
                        height=100,
                        key="kamp_yapistir_kodlar",
                        placeholder="Örnek:\n101\n102\n103\nveya: 101, 102, 103"
                    )
                    if st.button("🔍 Kodları Bul ve Seç", use_container_width=True):
                        if yapistir_text.strip():
                            # Kodları parse et (virgül, satır sonu, boşluk)
                            import re
                            kodlar = re.split(r'[,\n\s]+', yapistir_text.strip())
                            kodlar = [k.strip() for k in kodlar if k.strip()]

                            # Mağaza options'dan eşleşenleri bul
                            eslesenler = []
                            for opt in magaza_options:
                                opt_kod = opt.split(" - ")[0].strip()
                                if opt_kod in kodlar:
                                    eslesenler.append(opt)

                            if eslesenler:
                                # Hem multiselect key'ini hem de onay state'ini güncelle
                                st.session_state['kamp_magaza_select'] = eslesenler
                                st.session_state['kamp_secili_magazalar'] = eslesenler
                                st.session_state['kampanya_secili_magazalar'] = eslesenler  # Direkt onayla
                                # Cache temizle
                                if 'kamp_excel_bytes' in st.session_state:
                                    del st.session_state['kamp_excel_bytes']
                                if 'kampanya_sonuc' in st.session_state:
                                    del st.session_state['kampanya_sonuc']
                                st.success(f"✅ {len(eslesenler)}/{len(kodlar)} mağaza bulundu ve onaylandı!")
                                st.rerun()
                            else:
                                st.warning(f"⚠️ Girilen {len(kodlar)} koddan hiçbiri mevcut mağazalarla eşleşmedi")
                                st.info(f"Mevcut mağaza kodları: {', '.join([o.split(' - ')[0] for o in magaza_options[:5]])}...")

                secili_magazalar = st.multiselect(
                    "Mağaza Seçin (zorunlu):",
                    options=magaza_options,
                    default=[m for m in st.session_state.get('kamp_secili_magazalar', []) if m in magaza_options],
                    key="kamp_magaza_select",
                    help="Kampanya yapılacak mağazaları seçin veya yukarıdan yapıştırın"
                )

                # Seçim onay butonu
                col_onay, col_temizle = st.columns([3, 1])
                with col_onay:
                    if st.button("✅ Seçimi Onayla", type="primary", use_container_width=True, disabled=len(secili_magazalar) == 0):
                        st.session_state['kamp_secili_magazalar'] = secili_magazalar
                        st.session_state['kampanya_secili_magazalar'] = secili_magazalar
                        # Excel cache temizle
                        if 'kamp_excel_bytes' in st.session_state:
                            del st.session_state['kamp_excel_bytes']
                        if 'kampanya_sonuc' in st.session_state:
                            del st.session_state['kampanya_sonuc']

                with col_temizle:
                    if st.button("🔄 Temizle", use_container_width=True):
                        for key in ['kamp_secili_sm', 'kamp_secili_bs', 'kamp_secili_magazalar',
                                    'kampanya_secili_magazalar', 'kampanya_sonuc', 'kamp_excel_bytes']:
                            if key in st.session_state:
                                del st.session_state[key]
                        st.rerun()

                # Session state'den oku
                if st.session_state.get('kampanya_secili_magazalar'):
                    secili_magazalar_aktif = st.session_state['kampanya_secili_magazalar']
                    stok_df_aktif = st.session_state.get('kampanya_stok_df', stok_df)

                    st.success(f"✅ {len(secili_magazalar_aktif)} mağaza onaylandı - Analiz edebilirsiniz")

                    # Seçili mağaza kodlarını al
                    secili_magaza_kodlari = [m.split(" - ")[0].strip() for m in secili_magazalar_aktif]

                    # Veriyi filtrele (strip ile)
                    stok_df_aktif['Kod'] = stok_df_aktif['Kod'].astype(str).str.strip()
                    filtered_df = stok_df_aktif[stok_df_aktif['Kod'].isin(secili_magaza_kodlari)]

                    st.markdown("---")

                    # =============================================================================
                    # ADIM 3: ANALİZ VE ÖNERİ - OPTİMİZE (dict lookup)
                    # =============================================================================
                    st.markdown("### 3️⃣ Kampanya Önerisi")

                    # Minimum fiyat filtresi
                    min_fiyat = st.number_input(
                        "💰 Minimum Ürün Fiyatı (₺)",
                        min_value=0.0,
                        max_value=100000.0,
                        value=0.0,
                        step=10.0,
                        help="Sadece bu fiyatın üzerindeki ürünler analize dahil edilir"
                    )

                    if st.button("🚀 Analiz Et ve Öner", type="primary", use_container_width=True):
                        with st.spinner("🔄 Lift algoritması çalışıyor..."):
                            # Cached parquet'ten veri al (RAM dostu)
                            performans_df = load_performans_data()
                            urun_mal_grubu_map = st.session_state.get("urun_mal_grubu_map") or get_urun_mal_grubu_map(performans_df)

                            if performans_df is None:
                                st.error("❌ Performans verisi yüklenemedi! Önce 'Performans Verisini Yükle' butonuna tıklayın.")
                            else:
                                # ÖNEMLİ: Aggregasyonları DÖNGÜ DIŞINDA bir kere hazırla (CACHED!)
                                agg = prepare_lift_aggregations("cached")
                                bench_total = agg['bench_total']
                                store_totals = agg['store_totals']
                                store_sku_qty = agg['store_sku_qty']
                                bench_sku_qty = agg['bench_sku_qty']
                                store_grp_qty = agg['store_grp_qty']
                                bench_grp_qty = agg['bench_grp_qty']
                                store_ust_grp_qty = agg['store_ust_grp_qty']
                                bench_ust_grp_qty = agg['bench_ust_grp_qty']
                                urun_mal_grubu_agg = agg['urun_mal_grubu']
                                urun_ust_mal_grubu_agg = agg['urun_ust_mal_grubu']

                                eps = 0.0001
                                sonuclar = []

                                # Ürün kodlarını strip'le (bir kere)
                                filtered_df['Ürün Kodu'] = filtered_df['Ürün Kodu'].astype(str).str.strip()

                                for _, row in filtered_df.iterrows():
                                    magaza_kodu = str(row['Kod']).strip()
                                    urun_kodu = str(row['Ürün Kodu']).strip()
                                    urun_adi = row['Ürün Tanımı']
                                    stok = row.get('Stok', 0) or 0
                                    satis_fiyati = row.get('Satış Fiyatı', 0)
                                    alis_fiyati = row.get('Alış', 0)
                                    marj = row.get('Marj', 0)

                                    # Fiyatı temizle
                                    if isinstance(satis_fiyati, str):
                                        try:
                                            satis_fiyati = float(satis_fiyati.replace('₺', '').replace('.', '').replace(',', '.').strip())
                                        except:
                                            satis_fiyati = 0

                                    # Minimum fiyat filtresi - altındakileri atla
                                    if min_fiyat > 0 and satis_fiyati < min_fiyat:
                                        continue

                                    # Mal grubunu ve üst mal grubunu Excel'den al (zorunlu kolonlar)
                                    mal_grubu = str(row.get('Mal Grubu', '')).strip()
                                    ust_mal_grubu = str(row.get('Üst Mal Grubu', '')).strip()

                                    # Dict lookup ile hızlı lift hesaplama
                                    store_total = store_totals.get(magaza_kodu, 0)

                                    if store_total == 0:
                                        # Mağaza verisi yok - stok bazlı öneri
                                        lift = 1.0
                                        sku_trusted = False
                                        if stok >= 20:
                                            fit_score = 70
                                            neden = f"📦 Yüksek stok ({stok} adet) - Mağaza verisi yok"
                                        elif stok >= 10:
                                            fit_score = 55
                                            neden = f"📦 Orta stok ({stok} adet) - Mağaza verisi yok"
                                        elif stok >= 5:
                                            fit_score = 40
                                            neden = f"📦 Düşük stok ({stok} adet) - Mağaza verisi yok"
                                        else:
                                            fit_score = 30
                                            neden = f"➖ Az stok ({stok} adet) - Mağaza verisi yok"
                                    else:
                                        # === 3 SEVİYELİ LIFT HESAPLAMA ===

                                        # 1. SKU bazlı lift
                                        sku_qty = store_sku_qty.get((magaza_kodu, urun_kodu), 0)
                                        bench_qty = bench_sku_qty.get(urun_kodu, 0)

                                        store_share = (sku_qty / (store_total + eps)) * 100
                                        bench_share = (bench_qty / (bench_total + eps)) * 100

                                        lift = (store_share + eps) / (bench_share + eps)
                                        fit_sku = min(max((lift - 0.5) / 1.5, 0), 1) * 100

                                        # 2. Mal grubu bazlı lift
                                        grp_qty = store_grp_qty.get((magaza_kodu, mal_grubu), 0) if mal_grubu else 0
                                        grp_bench = bench_grp_qty.get(mal_grubu, 0) if mal_grubu else 0

                                        if grp_bench > 0:
                                            grp_share = (grp_qty / (store_total + eps)) * 100
                                            grp_bench_share = (grp_bench / (bench_total + eps)) * 100
                                            lift_grp = (grp_share + eps) / (grp_bench_share + eps)
                                            fit_grp = min(max((lift_grp - 0.5) / 1.5, 0), 1) * 100
                                        else:
                                            fit_grp = fit_sku

                                        # 3. Üst mal grubu bazlı lift
                                        ust_grp_qty = store_ust_grp_qty.get((magaza_kodu, ust_mal_grubu), 0) if ust_mal_grubu else 0
                                        ust_grp_bench = bench_ust_grp_qty.get(ust_mal_grubu, 0) if ust_mal_grubu else 0

                                        if ust_grp_bench > 0:
                                            ust_grp_share = (ust_grp_qty / (store_total + eps)) * 100
                                            ust_grp_bench_share = (ust_grp_bench / (bench_total + eps)) * 100
                                            lift_ust = (ust_grp_share + eps) / (ust_grp_bench_share + eps)
                                            fit_ust = min(max((lift_ust - 0.5) / 1.5, 0), 1) * 100
                                        else:
                                            fit_ust = fit_grp

                                        # === 3 SEVİYELİ HİERARŞİK BLEND ===
                                        alpha_sku = sku_qty / (sku_qty + 5)
                                        alpha_grp = grp_qty / (grp_qty + 20)

                                        fit_score = (
                                            alpha_sku * fit_sku +
                                            (1 - alpha_sku) * alpha_grp * fit_grp +
                                            (1 - alpha_sku) * (1 - alpha_grp) * fit_ust
                                        )

                                        sku_trusted = sku_qty >= 3 and bench_qty >= 30

                                        # Neden öneriliyor?
                                        if lift > 1.5:
                                            neden = f"🔥 Yüksek SKU lift ({lift:.1f}x) - Mağaza bu üründe güçlü"
                                        elif lift > 1.0:
                                            neden = f"✅ Pozitif SKU lift ({lift:.1f}x) - Ortalamanın üstünde"
                                        elif ust_mal_grubu and fit_ust > 60:
                                            neden = f"📊 {ust_mal_grubu} kategorisinde güçlü"
                                        elif mal_grubu and fit_grp > 60:
                                            neden = f"📈 {mal_grubu} grubunda güçlü"
                                        elif stok > 10:
                                            neden = f"📦 Yüksek stok ({stok} adet) - Eritilmeli"
                                        else:
                                            neden = f"➖ Standart performans (lift: {lift:.1f}x)"

                                    # Stok değeri hesapla
                                    stok_tl = stok * (satis_fiyati if isinstance(satis_fiyati, (int, float)) else 0)

                                    sonuclar.append({
                                        'SM': row.get('SM', ''),
                                        'BS': row.get('BS', ''),
                                        'Kod': magaza_kodu,
                                        'Mağaza Adı': row.get('Mağaza Adı', ''),
                                        'Ürün Kodu': urun_kodu,
                                        'Ürün Tanımı': urun_adi,
                                        'Stok': stok,
                                        'Alış': alis_fiyati,
                                        'Satış Fiyatı': satis_fiyati,
                                        'Marj': marj,
                                        'kdv': row.get('kdv', row.get('KDV', '')),
                                        'yeni fiyat': '',  # Kullanıcı dolduracak
                                        'yeni marj': '',  # Formül ile hesaplanacak
                                        'Stok TL': stok_tl,
                                        'üst mal grubu': row.get('Üst Mal Grubu', ''),
                                        'mal grubu': row.get('Mal Grubu', ''),
                                        'Lift Skoru': round(fit_score, 1),
                                        'Lift': round(lift, 2),
                                        'Öneri Nedeni': neden
                                    })

                                # DataFrame oluştur ve sırala
                                sonuc_df = pd.DataFrame(sonuclar)
                                # Kod kolonunu string'e çevir (pyarrow mixed type hatası için)
                                if 'Kod' in sonuc_df.columns:
                                    sonuc_df['Kod'] = sonuc_df['Kod'].astype(str).str.strip()

                                # Mağaza bazında grupla ve her mağaza için en iyi ürünleri seç
                                sonuc_df = sonuc_df.sort_values(
                                    ['Kod', 'Lift Skoru', 'Stok TL'],
                                    ascending=[True, False, False]
                                )

                                # Session state'e kaydet
                                st.session_state['kampanya_sonuc'] = sonuc_df

                                st.success(f"✅ Analiz tamamlandı! {len(sonuc_df)} ürün-mağaza kombinasyonu değerlendirildi.")

                    # Sonuçları göster
                    if 'kampanya_sonuc' in st.session_state:
                        sonuc_df = st.session_state['kampanya_sonuc']

                        st.markdown("---")
                        st.markdown("### 📊 Sonuçlar")

                        # Özet istatistikler
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Toplam Satır", f"{len(sonuc_df):,}")
                        with col2:
                            st.metric("Benzersiz Ürün", f"{sonuc_df['Ürün Kodu'].nunique():,}")
                        with col3:
                            st.metric("Toplam Stok", f"{sonuc_df['Stok'].sum():,}")
                        with col4:
                            st.metric("Toplam Stok TL", f"₺{sonuc_df['Stok TL'].sum():,.0f}")

                        # Filtreleme seçenekleri - FORM ile (rerun fırtınası önlenir)
                        st.markdown("#### 🎯 Sonuç Filtresi")
                        with st.form("sonuc_filtre_form"):
                            col_f1, col_f2 = st.columns(2)
                            with col_f1:
                                min_skor_input = st.slider("Minimum Lift Skoru", 0, 100,
                                    st.session_state.get('kamp_min_skor', 30), 5)
                            with col_f2:
                                min_stok_input = st.number_input("Minimum Stok", 0, 1000,
                                    st.session_state.get('kamp_min_stok', 1))
                            apply_filter = st.form_submit_button("🔍 Filtre Uygula", use_container_width=True)

                        if apply_filter:
                            st.session_state['kamp_min_skor'] = min_skor_input
                            st.session_state['kamp_min_stok'] = min_stok_input
                            # Excel cache'ini temizle (yeni filtre = yeni excel)
                            if 'kamp_excel_bytes' in st.session_state:
                                del st.session_state['kamp_excel_bytes']

                        min_skor = st.session_state.get('kamp_min_skor', 30)
                        min_stok = st.session_state.get('kamp_min_stok', 1)

                        # Filtrele
                        filtered_sonuc = sonuc_df[
                            (sonuc_df['Lift Skoru'] >= min_skor) &
                            (sonuc_df['Stok'] >= min_stok)
                        ]

                        st.info(f"📋 Filtrelenmiş: {len(filtered_sonuc):,} satır (Skor ≥ {min_skor}, Stok ≥ {min_stok})")

                        # Tablo göster (max 2000 satır UI için)
                        display_df = filtered_sonuc.head(2000) if len(filtered_sonuc) > 2000 else filtered_sonuc
                        if len(filtered_sonuc) > 2000:
                            st.warning(f"⚠️ Tabloda ilk 2000 satır gösteriliyor. Tamamı Excel'de.")
                        st.dataframe(
                            display_df,
                            use_container_width=True,
                            height=400
                        )

                        # Excel indirme - SADECE BUTONA BASINCA OLUŞTUR
                        st.markdown("---")
                        st.markdown("### 📥 Excel İndir")

                        col_prep, col_dl = st.columns([1, 2])
                        with col_prep:
                            if st.button("📦 Excel Hazırla", type="secondary", use_container_width=True):
                                with st.spinner("Excel hazırlanıyor..."):
                                    st.session_state['kamp_excel_bytes'] = write_excel_with_formulas(
                                        filtered_sonuc, sheet_name='Kampanya Önerisi'
                                    ).getvalue()
                                st.success("✅ Excel hazır!")

                        with col_dl:
                            excel_ready = 'kamp_excel_bytes' in st.session_state
                            st.download_button(
                                label="📥 Kampanya Önerisini İndir" if excel_ready else "📥 Önce Excel Hazırla",
                                data=st.session_state.get('kamp_excel_bytes', b''),
                                file_name=f"kampanya_onerisi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True,
                                disabled=not excel_ready
                            )

                        st.caption("💡 'Excel Hazırla' → 'İndir' → 'yeni fiyat' doldur → 'yeni marj' otomatik hesaplanır.")

                else:
                    st.warning("👆 En az bir mağaza seçin.")

        except Exception as e:
            st.error(f"❌ Excel okuma hatası: {str(e)}")
            st.info("Dosya formatını kontrol edin. Beklenen kolonlar: SM, BS, Kod, Mağaza Adı, Ürün Kodu, Ürün Tanımı, Stok, Alış, Satış Fiyatı...")

    else:
        st.info("👆 Excel dosyası yükleyin.")

# =============================================================================
# WHATSAPP KANALI KAMPANYA MODU
# =============================================================================
elif mod_secim == "📱 WhatsApp Kanalı Kampanya":
    st.markdown("""
    <div class="secim-rehberi">
        <strong>📱 WhatsApp Kanalı Kampanya Nasıl Çalışır?</strong><br>
        1. Stok Excel dosyasını yükleyin<br>
        2. Sistem otomatik olarak WhatsApp kanalındaki mağazaları filtreler<br>
        3. Lift algoritmasıyla kampanya önerisi oluşturur<br>
        4. Excel olarak indirin ve fiyatları belirleyin
    </div>
    """, unsafe_allow_html=True)

    # WhatsApp kanalındaki mağaza kodları
    whatsapp_magaza_kodlari = list(MAGAZALAR.keys())
    st.info(f"📱 WhatsApp kanalında {len(whatsapp_magaza_kodlari)} mağaza var")

    # Excel yükleme
    st.markdown("### 1️⃣ Stok Verisini Yükleyin")

    uploaded_file_wp = st.file_uploader(
        "Stok Excel dosyasını yükleyin (.xlsx)",
        type=['xlsx'],
        key="stok_excel_upload_wp",
        help="Kolonlar: Kod, Mağaza Adı, Ürün Kodu, Ürün Tanımı, Stok, Satış Fiyatı..."
    )

    if uploaded_file_wp is not None:
        try:
            stok_df = pd.read_excel(uploaded_file_wp)
            stok_df.columns = stok_df.columns.str.strip()

            # Kolon isimlerini standartlaştır (case-insensitive)
            kolon_mapping = {
                'üst mal grubu': 'Üst Mal Grubu',
                'mal grubu': 'Mal Grubu',
                'ürün kodu': 'Ürün Kodu',
                'ürün tanımı': 'Ürün Tanımı',
                'mağaza adı': 'Mağaza Adı',
                'satış fiyatı': 'Satış Fiyatı',
                'kod': 'Kod',
                'stok': 'Stok'
            }
            stok_df.columns = [kolon_mapping.get(col.lower(), col) for col in stok_df.columns]

            gerekli_kolonlar = ['Kod', 'Mağaza Adı', 'Ürün Kodu', 'Ürün Tanımı', 'Stok', 'Satış Fiyatı', 'Üst Mal Grubu', 'Mal Grubu']
            eksik_kolonlar = [k for k in gerekli_kolonlar if k not in stok_df.columns]

            if eksik_kolonlar:
                st.error(f"❌ Eksik kolonlar: {', '.join(eksik_kolonlar)}")
            else:
                # WhatsApp mağazalarını filtrele
                stok_df['Kod'] = stok_df['Kod'].astype(str).str.strip()
                filtered_df = stok_df[stok_df['Kod'].isin(whatsapp_magaza_kodlari)]

                if len(filtered_df) == 0:
                    st.warning("⚠️ Excel'de WhatsApp kanalı mağazalarından hiçbiri bulunamadı!")
                    st.info(f"Beklenen mağaza kodları: {', '.join(whatsapp_magaza_kodlari[:5])}...")
                else:
                    bulunan_magazalar = filtered_df['Kod'].nunique()
                    st.success(f"✅ {len(filtered_df):,} satır yüklendi ({bulunan_magazalar}/{len(whatsapp_magaza_kodlari)} mağaza bulundu)")

                    with st.expander("📋 Veri Önizleme"):
                        st.dataframe(filtered_df.head(10))

                    st.markdown("---")
                    st.markdown("### 2️⃣ Kampanya Önerisi")

                    if st.button("🚀 Analiz Et ve Öner", type="primary", use_container_width=True, key="analiz_wp"):
                        with st.spinner("🔄 Lift algoritması çalışıyor..."):
                            # Cached parquet'ten veri al (RAM dostu)
                            performans_df = load_performans_data()
                            urun_mal_grubu_map = st.session_state.get("urun_mal_grubu_map") or get_urun_mal_grubu_map(performans_df)

                            if performans_df is None:
                                st.error("❌ Performans verisi yüklenemedi! Önce 'Performans Verisini Yükle' butonuna tıklayın.")
                            else:
                                # CACHED aggregasyonlar - RAM dostu
                                agg = prepare_lift_aggregations("cached")
                                bench_total = agg['bench_total']
                                store_totals = agg['store_totals']
                                store_sku_qty = agg['store_sku_qty']
                                bench_sku_qty = agg['bench_sku_qty']
                                store_grp_qty = agg['store_grp_qty']
                                bench_grp_qty = agg['bench_grp_qty']
                                store_ust_grp_qty = agg['store_ust_grp_qty']
                                bench_ust_grp_qty = agg['bench_ust_grp_qty']
                                urun_mal_grubu_agg = agg['urun_mal_grubu']
                                urun_ust_mal_grubu_agg = agg['urun_ust_mal_grubu']

                                eps = 0.0001
                                sonuclar = []

                                filtered_df['Ürün Kodu'] = filtered_df['Ürün Kodu'].astype(str).str.strip()

                                for _, row in filtered_df.iterrows():
                                    magaza_kodu = str(row['Kod']).strip()
                                    urun_kodu = str(row['Ürün Kodu']).strip()
                                    urun_adi = row['Ürün Tanımı']
                                    stok = row.get('Stok', 0) or 0
                                    satis_fiyati = row.get('Satış Fiyatı', 0)
                                    alis_fiyati = row.get('Alış', 0)
                                    marj = row.get('Marj', 0)

                                    if isinstance(satis_fiyati, str):
                                        try:
                                            satis_fiyati = float(satis_fiyati.replace('₺', '').replace('.', '').replace(',', '.').strip())
                                        except:
                                            satis_fiyati = 0

                                    # Mal grubunu ve üst mal grubunu Excel'den al (zorunlu kolonlar)
                                    mal_grubu = str(row.get('Mal Grubu', '')).strip()
                                    ust_mal_grubu = str(row.get('Üst Mal Grubu', '')).strip()

                                    store_total = store_totals.get(magaza_kodu, 0)

                                    if store_total == 0:
                                        lift = 1.0
                                        sku_trusted = False
                                        if stok >= 20:
                                            fit_score = 70
                                            neden = f"📦 Yüksek stok ({stok} adet) - Mağaza verisi yok"
                                        elif stok >= 10:
                                            fit_score = 55
                                            neden = f"📦 Orta stok ({stok} adet) - Mağaza verisi yok"
                                        elif stok >= 5:
                                            fit_score = 40
                                            neden = f"📦 Düşük stok ({stok} adet) - Mağaza verisi yok"
                                        else:
                                            fit_score = 30
                                            neden = f"➖ Az stok ({stok} adet) - Mağaza verisi yok"
                                    else:
                                        # === 3 SEVİYELİ LIFT HESAPLAMA ===

                                        # 1. SKU bazlı lift
                                        sku_qty = store_sku_qty.get((magaza_kodu, urun_kodu), 0)
                                        bench_qty = bench_sku_qty.get(urun_kodu, 0)

                                        store_share = (sku_qty / (store_total + eps)) * 100
                                        bench_share = (bench_qty / (bench_total + eps)) * 100

                                        lift = (store_share + eps) / (bench_share + eps)
                                        fit_sku = min(max((lift - 0.5) / 1.5, 0), 1) * 100

                                        # 2. Mal grubu bazlı lift
                                        grp_qty = store_grp_qty.get((magaza_kodu, mal_grubu), 0) if mal_grubu else 0
                                        grp_bench = bench_grp_qty.get(mal_grubu, 0) if mal_grubu else 0

                                        if grp_bench > 0:
                                            grp_share = (grp_qty / (store_total + eps)) * 100
                                            grp_bench_share = (grp_bench / (bench_total + eps)) * 100
                                            lift_grp = (grp_share + eps) / (grp_bench_share + eps)
                                            fit_grp = min(max((lift_grp - 0.5) / 1.5, 0), 1) * 100
                                        else:
                                            fit_grp = fit_sku

                                        # 3. Üst mal grubu bazlı lift
                                        ust_grp_qty = store_ust_grp_qty.get((magaza_kodu, ust_mal_grubu), 0) if ust_mal_grubu else 0
                                        ust_grp_bench = bench_ust_grp_qty.get(ust_mal_grubu, 0) if ust_mal_grubu else 0

                                        if ust_grp_bench > 0:
                                            ust_grp_share = (ust_grp_qty / (store_total + eps)) * 100
                                            ust_grp_bench_share = (ust_grp_bench / (bench_total + eps)) * 100
                                            lift_ust = (ust_grp_share + eps) / (ust_grp_bench_share + eps)
                                            fit_ust = min(max((lift_ust - 0.5) / 1.5, 0), 1) * 100
                                        else:
                                            fit_ust = fit_grp

                                        # === 3 SEVİYELİ HİERARŞİK BLEND ===
                                        alpha_sku = sku_qty / (sku_qty + 5)
                                        alpha_grp = grp_qty / (grp_qty + 20)

                                        fit_score = (
                                            alpha_sku * fit_sku +
                                            (1 - alpha_sku) * alpha_grp * fit_grp +
                                            (1 - alpha_sku) * (1 - alpha_grp) * fit_ust
                                        )

                                        sku_trusted = sku_qty >= 3 and bench_qty >= 30

                                        # Neden öneriliyor?
                                        if lift > 1.5:
                                            neden = f"🔥 Yüksek SKU lift ({lift:.1f}x) - Mağaza bu üründe güçlü"
                                        elif lift > 1.0:
                                            neden = f"✅ Pozitif SKU lift ({lift:.1f}x) - Ortalamanın üstünde"
                                        elif ust_mal_grubu and fit_ust > 60:
                                            neden = f"📊 {ust_mal_grubu} kategorisinde güçlü"
                                        elif mal_grubu and fit_grp > 60:
                                            neden = f"📈 {mal_grubu} grubunda güçlü"
                                        elif stok > 10:
                                            neden = f"📦 Yüksek stok ({stok} adet) - Eritilmeli"
                                        else:
                                            neden = f"➖ Standart performans (lift: {lift:.1f}x)"

                                    stok_tl = stok * (satis_fiyati if isinstance(satis_fiyati, (int, float)) else 0)
                                    magaza_adi = MAGAZALAR.get(magaza_kodu, row.get('Mağaza Adı', ''))

                                    sonuclar.append({
                                        'SM': row.get('SM', ''),
                                        'BS': row.get('BS', ''),
                                        'Kod': magaza_kodu,
                                        'Mağaza Adı': magaza_adi,
                                        'Ürün Kodu': urun_kodu,
                                        'Ürün Tanımı': urun_adi,
                                        'Stok': stok,
                                        'Alış': alis_fiyati,
                                        'Satış Fiyatı': satis_fiyati,
                                        'Marj': marj,
                                        'kdv': row.get('kdv', row.get('KDV', '')),
                                        'yeni fiyat': '',
                                        'yeni marj': '',  # Formül ile hesaplanacak
                                        'Stok TL': stok_tl,
                                        'üst mal grubu': row.get('Üst Mal Grubu', ''),
                                        'mal grubu': row.get('Mal Grubu', ''),
                                        'Lift Skoru': round(fit_score, 1),
                                        'Lift': round(lift, 2),
                                        'Öneri Nedeni': neden
                                    })

                                sonuc_df = pd.DataFrame(sonuclar)
                                # Kod kolonunu string'e çevir (pyarrow mixed type hatası için)
                                if 'Kod' in sonuc_df.columns:
                                    sonuc_df['Kod'] = sonuc_df['Kod'].astype(str).str.strip()
                                sonuc_df = sonuc_df.sort_values(['Kod', 'Lift Skoru', 'Stok TL'], ascending=[True, False, False])
                                st.session_state['wp_kampanya_sonuc'] = sonuc_df
                                st.success(f"✅ Analiz tamamlandı! {len(sonuc_df)} ürün-mağaza kombinasyonu")

                    # Sonuçları göster
                    if 'wp_kampanya_sonuc' in st.session_state:
                        sonuc_df = st.session_state['wp_kampanya_sonuc']

                        st.markdown("---")
                        st.markdown("### 📊 Sonuçlar")

                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Toplam Satır", f"{len(sonuc_df):,}")
                        with col2:
                            st.metric("Benzersiz Ürün", f"{sonuc_df['Ürün Kodu'].nunique():,}")
                        with col3:
                            st.metric("Toplam Stok", f"{sonuc_df['Stok'].sum():,}")
                        with col4:
                            st.metric("Toplam Stok TL", f"₺{sonuc_df['Stok TL'].sum():,.0f}")

                        # Filtreleme seçenekleri - FORM ile (rerun fırtınası önlenir)
                        st.markdown("#### 🎯 Sonuç Filtresi")
                        with st.form("wp_sonuc_filtre_form"):
                            col_f1, col_f2 = st.columns(2)
                            with col_f1:
                                min_skor_input = st.slider("Minimum Lift Skoru", 0, 100,
                                    st.session_state.get('wp_min_skor', 30), 5)
                            with col_f2:
                                min_stok_input = st.number_input("Minimum Stok", 0, 1000,
                                    st.session_state.get('wp_min_stok', 1))
                            apply_filter = st.form_submit_button("🔍 Filtre Uygula", use_container_width=True)

                        if apply_filter:
                            st.session_state['wp_min_skor'] = min_skor_input
                            st.session_state['wp_min_stok'] = min_stok_input
                            # Excel cache'ini temizle
                            if 'wp_excel_bytes' in st.session_state:
                                del st.session_state['wp_excel_bytes']

                        min_skor = st.session_state.get('wp_min_skor', 30)
                        min_stok = st.session_state.get('wp_min_stok', 1)

                        filtered_sonuc = sonuc_df[
                            (sonuc_df['Lift Skoru'] >= min_skor) &
                            (sonuc_df['Stok'] >= min_stok)
                        ]

                        st.info(f"📋 Filtrelenmiş: {len(filtered_sonuc):,} satır (Skor ≥ {min_skor}, Stok ≥ {min_stok})")

                        # Tablo göster (max 2000 satır UI için)
                        display_df = filtered_sonuc.head(2000) if len(filtered_sonuc) > 2000 else filtered_sonuc
                        if len(filtered_sonuc) > 2000:
                            st.warning(f"⚠️ Tabloda ilk 2000 satır gösteriliyor. Tamamı Excel'de.")
                        st.dataframe(display_df, use_container_width=True, height=400)

                        # Excel indirme - SADECE BUTONA BASINCA OLUŞTUR
                        st.markdown("---")
                        st.markdown("### 📥 Excel İndir")

                        col_prep, col_dl = st.columns([1, 2])
                        with col_prep:
                            if st.button("📦 Excel Hazırla", type="secondary", use_container_width=True, key="wp_excel_prep"):
                                with st.spinner("Excel hazırlanıyor..."):
                                    st.session_state['wp_excel_bytes'] = write_excel_with_formulas(
                                        filtered_sonuc, sheet_name='WhatsApp Kampanya'
                                    ).getvalue()
                                st.success("✅ Excel hazır!")

                        with col_dl:
                            excel_ready = 'wp_excel_bytes' in st.session_state
                            st.download_button(
                                label="📥 WhatsApp Kampanya İndir" if excel_ready else "📥 Önce Excel Hazırla",
                                data=st.session_state.get('wp_excel_bytes', b''),
                                file_name=f"whatsapp_kampanya_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True,
                                disabled=not excel_ready
                            )

                        st.caption("💡 'Excel Hazırla' → 'İndir' → 'yeni fiyat' doldur → 'yeni marj' otomatik hesaplanır.")

        except Exception as e:
            st.error(f"❌ Excel okuma hatası: {str(e)}")
    else:
        st.info("👆 Excel dosyası yükleyin.")

# =============================================================================
# TOPLU MESAJ MODU
# =============================================================================
elif mod_secim == "📤 Toplu Mesaj":
    st.markdown("""
    <div class="secim-rehberi">
        <strong>📤 Toplu Mesaj Nasıl Çalışır?</strong><br>
        1. İndirim yapılacak ürünlerin olduğu Excel'i yükleyin<br>
        2. Kampanya bitiş tarihini seçin<br>
        3. Her mağaza için otomatik mesaj oluşturulur
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 1️⃣ Excel Yükle")
    uploaded_toplu = st.file_uploader(
        "İndirimli ürün listesini yükleyin",
        type=['xlsx', 'xls'],
        key="toplu_excel"
    )

    st.markdown("### 2️⃣ Kampanya Bitiş Tarihi")
    from datetime import date, timedelta
    bitis_tarihi = st.date_input(
        "Kampanya son günü:",
        value=date.today() + timedelta(days=1),
        key="toplu_bitis"
    )

    if uploaded_toplu:
        try:
            df_toplu = pd.read_excel(uploaded_toplu)

            # Gerekli kolonları kontrol et
            required_cols = ['Kod', 'Mağaza Adı', 'Ürün Tanımı', 'Satış Fiyatı', 'yeni fiyat']
            missing_cols = [c for c in required_cols if c not in df_toplu.columns]

            if missing_cols:
                st.error(f"❌ Eksik kolonlar: {', '.join(missing_cols)}")
                st.info("Beklenen kolonlar: " + ", ".join(required_cols))
            else:
                # Mağazalara göre grupla
                df_toplu['Kod'] = df_toplu['Kod'].astype(str).str.strip()
                magazalar_grouped = df_toplu.groupby('Kod')

                st.success(f"✅ {len(df_toplu)} ürün, {len(magazalar_grouped)} mağaza bulundu")

                st.markdown("---")
                st.markdown("### 3️⃣ Mesajları Oluştur")

                if st.button("🚀 Tüm Mesajları Oluştur", type="primary", use_container_width=True):
                    bitis_str = bitis_tarihi.strftime("%d.%m.%Y")

                    # Her mağaza için mesaj oluştur
                    all_messages = {}

                    for magaza_kodu, group in magazalar_grouped:
                        magaza_adi = group['Mağaza Adı'].iloc[0]
                        urunler = []

                        for _, row in group.iterrows():
                            urun_adi = str(row['Ürün Tanımı']).strip()
                            eski_fiyat = float(row['Satış Fiyatı']) if pd.notna(row['Satış Fiyatı']) else 0
                            yeni_fiyat = float(row['yeni fiyat']) if pd.notna(row['yeni fiyat']) else 0

                            if eski_fiyat > 0 and yeni_fiyat > 0:
                                indirim = round((eski_fiyat - yeni_fiyat) / eski_fiyat * 100)
                                emoji = get_emoji(urun_adi)
                                urunler.append({
                                    'ad': urun_adi,
                                    'eski': eski_fiyat,
                                    'yeni': yeni_fiyat,
                                    'indirim': indirim,
                                    'emoji': emoji
                                })

                        if urunler:
                            # Mesaj oluştur
                            mesaj_lines = [
                                "WHATSAPP TAKİPÇİLERİNE ÖZEL",
                                f"🛒 A101 {magaza_adi}",
                                "",
                                f"🔥 YARIN {len(urunler)} Üründe %50 İNDİRİM var!",
                                "",
                                "İNDİRİM KODUNUZ: ANTALYA101 (KODU KASİYERE SÖYLEMENİZ YETERLİ)",
                                ""
                            ]

                            for u in urunler:
                                mesaj_lines.append(f"{u['emoji']} {u['ad'].upper()}")
                                mesaj_lines.append(f"✅ {u['yeni']:,.2f}₺ | Eski: {u['eski']:,.2f}₺ (%{u['indirim']} İNDİRİM)")
                                mesaj_lines.append("")

                            mesaj_lines.append(f"📅 Son gün: {bitis_str} | 📍 Stoklarla sınırlıdır.")

                            all_messages[magaza_kodu] = {
                                'ad': magaza_adi,
                                'mesaj': "\n".join(mesaj_lines),
                                'urun_sayisi': len(urunler)
                            }

                    # Session state'e kaydet
                    st.session_state['toplu_mesajlar'] = all_messages
                    st.rerun()

                # Mesajları göster
                if 'toplu_mesajlar' in st.session_state and st.session_state['toplu_mesajlar']:
                    mesajlar = st.session_state['toplu_mesajlar']

                    st.markdown("---")
                    st.markdown(f"### 📬 {len(mesajlar)} Mağaza İçin Mesajlar")

                    for magaza_kodu, data in mesajlar.items():
                        with st.expander(f"🏪 {magaza_kodu} - {data['ad']} ({data['urun_sayisi']} ürün)", expanded=False):
                            st.markdown(f"""
                            <div class="mesaj-onizleme">
{data['mesaj']}
                            </div>
                            """, unsafe_allow_html=True)

                            # Kopyala ve Düzenle butonları
                            col_copy, col_edit = st.columns(2)
                            with col_copy:
                                if st.button(f"📋 Kopyala", key=f"copy_{magaza_kodu}", use_container_width=True):
                                    st.code(data['mesaj'], language=None)
                                    st.success("👆 Yukarıdaki metni seçip kopyalayın")
                            with col_edit:
                                if st.button(f"✏️ Düzenle", key=f"edit_{magaza_kodu}", use_container_width=True):
                                    st.session_state[f'editing_{magaza_kodu}'] = True

                            # Düzenleme modu
                            if st.session_state.get(f'editing_{magaza_kodu}', False):
                                edited_mesaj = st.text_area(
                                    "Mesajı düzenleyin:",
                                    value=data['mesaj'],
                                    height=300,
                                    key=f"textarea_{magaza_kodu}"
                                )
                                if st.button("💾 Kaydet", key=f"save_{magaza_kodu}"):
                                    st.session_state['toplu_mesajlar'][magaza_kodu]['mesaj'] = edited_mesaj
                                    st.session_state[f'editing_{magaza_kodu}'] = False
                                    st.rerun()

        except Exception as e:
            st.error(f"❌ Excel okuma hatası: {str(e)}")
    else:
        st.info("👆 Excel dosyası yükleyin.")

# Footer
st.markdown("---")
st.markdown("""
<p style="text-align:center; color:#888; font-size:12px;">
    A101 Kampanya Asistanı v4.0 - Mesaj Oluşturucu + Kampanya Oluşturucu + Toplu Mesaj<br>
    Yeni Mağazacılık A.Ş. © 2025
</p>
""", unsafe_allow_html=True)
